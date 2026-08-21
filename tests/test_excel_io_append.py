import openpyxl
import pandas as pd
from openpyxl.styles import Font

from core.excel_io import (
    append_leads, guess_target_field_mapping, find_header_row, read_sheet_headers, route_leads_by_cid,
)
from core.models import FieldMapping, LeadTemplateTab


def _make_accumulated_workbook(path: str) -> None:
    wb = openpyxl.Workbook()
    lookup = wb.active
    lookup.title = "Lookup"
    lookup.append(["CID", "Name"])
    lookup.append([100, "Campaign A"])
    lookup.append([200, "Campaign B"])

    accumulated = wb.create_sheet("Accumulated")
    accumulated.append(["Date", "CID", "Campaign Name", "Comment", "emailaddress", "firstname", "lastname", "company"])
    accumulated.append(["2026-01-01", 100, "=VLOOKUP(B2,Lookup!A:B,2,0)", "Delivered", "a@x.com", "A", "B", "X"])

    refund = wb.create_sheet("Refund")
    refund.append(["Date", "CID", "Campaign Name", "Comment", "emailaddress", "firstname", "lastname", "company", "Reason"])

    wb.save(path)


def _field_mapping() -> FieldMapping:
    return FieldMapping(email="emailaddress", first_name="firstname", last_name="lastname",
                         company="company", cid="CID")


def test_append_leads_to_accumulated_fills_by_header_and_shifts_formula(tmp_path):
    path = str(tmp_path / "accumulated.xlsx")
    _make_accumulated_workbook(path)

    leads_df = pd.DataFrame([
        {"CID": 200, "emailaddress": "c@y.com", "firstname": "C", "lastname": "D", "company": "Y"},
    ])

    append_leads(path, "Accumulated", leads_df, _field_mapping(), run_date="2026-08-08")

    wb = openpyxl.load_workbook(path)
    ws = wb["Accumulated"]

    assert ws.max_row == 3
    new_row = {cell.value for cell in ws[1]}
    assert new_row == {"Date", "CID", "Campaign Name", "Comment", "emailaddress", "firstname", "lastname", "company"}

    values = {ws.cell(row=1, column=c).value: ws.cell(row=3, column=c).value for c in range(1, 9)}
    assert values["Date"] == "2026-08-08"
    assert values["CID"] == 200
    assert values["Campaign Name"] == "=VLOOKUP(B3,Lookup!A:B,2,0)"
    assert values["Comment"] is None
    assert values["emailaddress"] == "c@y.com"
    assert values["firstname"] == "C"
    assert values["lastname"] == "D"
    assert values["company"] == "Y"


def test_append_leads_to_refund_fills_reason_column(tmp_path):
    path = str(tmp_path / "accumulated.xlsx")
    _make_accumulated_workbook(path)

    leads_df = pd.DataFrame([
        {"CID": 200, "emailaddress": "c@y.com", "firstname": "C", "lastname": "D", "company": "Y"},
    ], index=[42])

    append_leads(path, "Refund", leads_df, _field_mapping(), run_date="2026-08-08",
                 reasons={42: "Exclusion - domain; TAL - not found"})

    wb = openpyxl.load_workbook(path)
    ws = wb["Refund"]
    values = {ws.cell(row=1, column=c).value: ws.cell(row=2, column=c).value for c in range(1, 10)}
    assert values["Reason"] == "Exclusion - domain; TAL - not found"
    assert values["CID"] == 200


def test_append_leads_recognizes_refund_reason_header(tmp_path):
    path = str(tmp_path / "accumulated.xlsx")
    wb = openpyxl.Workbook()
    wb.active.title = "Lookup"
    refund = wb.create_sheet("Refund")
    refund.append(["Date", "CID", "emailaddress", "firstname", "lastname", "company", "Refund Reason"])
    wb.save(path)

    leads_df = pd.DataFrame([
        {"CID": 200, "emailaddress": "c@y.com", "firstname": "C", "lastname": "D", "company": "Y"},
    ], index=[7])

    append_leads(path, "Refund", leads_df, _field_mapping(), run_date="2026-08-08",
                 reasons={7: "Exclusion - domain"})

    wb2 = openpyxl.load_workbook(path)
    ws = wb2["Refund"]
    values = {ws.cell(row=1, column=c).value: ws.cell(row=2, column=c).value for c in range(1, 8)}
    assert values["Refund Reason"] == "Exclusion - domain"
    assert ws.max_column == 7  # no duplicate "Reason" column added


def test_append_leads_auto_adds_refund_reason_column_when_missing(tmp_path):
    path = str(tmp_path / "accumulated.xlsx")
    wb = openpyxl.Workbook()
    wb.active.title = "Lookup"
    refund = wb.create_sheet("Refund")
    refund.append(["Date", "CID", "emailaddress", "firstname", "lastname", "company"])
    wb.save(path)

    leads_df = pd.DataFrame([
        {"CID": 200, "emailaddress": "c@y.com", "firstname": "C", "lastname": "D", "company": "Y"},
    ], index=[3])

    append_leads(path, "Refund", leads_df, _field_mapping(), run_date="2026-08-08",
                 reasons={3: "TAL - not found"})

    wb2 = openpyxl.load_workbook(path)
    ws = wb2["Refund"]
    assert ws.cell(row=1, column=7).value == "Refund Reason"
    assert ws.cell(row=2, column=7).value == "TAL - not found"


def test_append_leads_first_batch_uses_row_two_formatting(tmp_path):
    path = str(tmp_path / "accumulated.xlsx")
    wb = openpyxl.Workbook()
    wb.active.title = "Lookup"
    accumulated = wb.create_sheet("Accumulated")
    accumulated.append(["CID", "emailaddress", "firstname", "lastname", "company"])
    template_font = Font(bold=True, color="FF0000FF")
    for col in range(1, 6):
        accumulated.cell(row=2, column=col).font = template_font
    wb.save(path)

    leads_df = pd.DataFrame([
        {"CID": 200, "emailaddress": "c@y.com", "firstname": "C", "lastname": "D", "company": "Y"},
    ])

    append_leads(path, "Accumulated", leads_df, _field_mapping(), run_date="2026-08-08")

    wb2 = openpyxl.load_workbook(path)
    ws = wb2["Accumulated"]
    assert ws.max_row == 2
    assert ws.cell(row=2, column=1).value == 200
    assert ws.cell(row=2, column=1).font.bold is True
    assert ws.cell(row=2, column=1).font.color.rgb == "FF0000FF"


def test_append_leads_subsequent_batch_uses_last_row_formatting(tmp_path):
    path = str(tmp_path / "accumulated.xlsx")
    wb = openpyxl.Workbook()
    wb.active.title = "Lookup"
    accumulated = wb.create_sheet("Accumulated")
    accumulated.append(["CID", "emailaddress", "firstname", "lastname", "company"])
    existing_font = Font(bold=True, color="FF00FF00")
    accumulated.append([100, "a@x.com", "A", "B", "X"])
    for col in range(1, 6):
        accumulated.cell(row=2, column=col).font = existing_font
    wb.save(path)

    leads_df = pd.DataFrame([
        {"CID": 200, "emailaddress": "c@y.com", "firstname": "C", "lastname": "D", "company": "Y"},
    ])

    append_leads(path, "Accumulated", leads_df, _field_mapping(), run_date="2026-08-08")

    wb2 = openpyxl.load_workbook(path)
    ws = wb2["Accumulated"]
    assert ws.max_row == 3
    assert ws.cell(row=3, column=1).value == 200
    assert ws.cell(row=3, column=1).font.bold is True
    assert ws.cell(row=3, column=1).font.color.rgb == "FF00FF00"


def test_append_leads_clear_existing_removes_old_rows_but_keeps_formatting(tmp_path):
    path = str(tmp_path / "lead_report.xlsx")
    wb = openpyxl.Workbook()
    wb.active.title = "Lookup"
    template = wb.create_sheet("Report")
    template.append(["CID", "emailaddress", "firstname", "lastname", "company"])
    existing_font = Font(bold=True, color="FF00FF00")
    template.append([100, "old@x.com", "Old", "Lead", "X"])
    template.append([101, "old2@x.com", "Old2", "Lead2", "X"])
    for row in (2, 3):
        for col in range(1, 6):
            template.cell(row=row, column=col).font = existing_font
    wb.save(path)

    leads_df = pd.DataFrame([
        {"CID": 200, "emailaddress": "new@y.com", "firstname": "New", "lastname": "Lead", "company": "Y"},
    ])

    append_leads(path, "Report", leads_df, _field_mapping(), run_date="2026-08-08", clear_existing=True)

    wb2 = openpyxl.load_workbook(path)
    ws = wb2["Report"]
    assert ws.max_row == 2  # header + exactly the one new lead, old rows gone
    assert ws.cell(row=2, column=2).value == "new@y.com"
    # Formatting from the removed old row 2 is preserved on the new row.
    assert ws.cell(row=2, column=1).font.bold is True
    assert ws.cell(row=2, column=1).font.color.rgb == "FF00FF00"


def test_append_leads_highlight_fill_colors_only_the_new_rows(tmp_path):
    path = str(tmp_path / "lead_report.xlsx")
    wb = openpyxl.Workbook()
    wb.active.title = "Lookup"
    template = wb.create_sheet("Report")
    template.append(["CID", "emailaddress", "firstname", "lastname", "company"])
    template.append([100, "old@x.com", "Old", "Lead", "X"])
    wb.save(path)

    leads_df = pd.DataFrame([
        {"CID": 200, "emailaddress": "new@y.com", "firstname": "New", "lastname": "Lead", "company": "Y"},
    ])

    append_leads(path, "Report", leads_df, _field_mapping(), run_date="2026-08-08", highlight_fill="C6E0B4")

    ws = openpyxl.load_workbook(path)["Report"]
    assert ws.cell(row=3, column=1).fill.fgColor.rgb == "00C6E0B4"
    # The pre-existing row must not have been touched.
    assert ws.cell(row=2, column=1).fill.fill_type is None


def test_append_leads_highlight_fill_clears_previous_run_highlight(tmp_path):
    path = str(tmp_path / "lead_report.xlsx")
    wb = openpyxl.Workbook()
    wb.active.title = "Lookup"
    template = wb.create_sheet("Report")
    template.append(["CID", "emailaddress", "firstname", "lastname", "company"])
    wb.save(path)

    # Simulate an earlier run that highlighted its own new row (row 2).
    highlighted_previously = pd.DataFrame([
        {"CID": 100, "emailaddress": "prev@x.com", "firstname": "Prev", "lastname": "Lead", "company": "X"},
    ])
    append_leads(path, "Report", highlighted_previously, _field_mapping(), run_date="2026-08-07",
                 highlight_fill="C6E0B4")

    new_leads_df = pd.DataFrame([
        {"CID": 200, "emailaddress": "new@y.com", "firstname": "New", "lastname": "Lead", "company": "Y"},
    ])
    append_leads(path, "Report", new_leads_df, _field_mapping(), run_date="2026-08-08", highlight_fill="C6E0B4")

    ws = openpyxl.load_workbook(path)["Report"]
    # Row 2 was highlighted by the first call — the second call must clear it.
    assert ws.cell(row=2, column=1).fill.fill_type is None
    # Row 3 (this run's new lead) is now the one highlighted.
    assert ws.cell(row=3, column=1).fill.fgColor.rgb == "00C6E0B4"


def test_append_leads_without_highlight_fill_leaves_fill_untouched(tmp_path):
    path = str(tmp_path / "lead_report.xlsx")
    wb = openpyxl.Workbook()
    wb.active.title = "Lookup"
    template = wb.create_sheet("Report")
    template.append(["CID", "emailaddress", "firstname", "lastname", "company"])
    wb.save(path)

    leads_df = pd.DataFrame([
        {"CID": 200, "emailaddress": "new@y.com", "firstname": "New", "lastname": "Lead", "company": "Y"},
    ])
    append_leads(path, "Report", leads_df, _field_mapping(), run_date="2026-08-08")

    ws = openpyxl.load_workbook(path)["Report"]
    assert ws.cell(row=2, column=1).fill.fill_type is None


def test_append_leads_clear_existing_on_empty_sheet_is_a_no_op(tmp_path):
    path = str(tmp_path / "lead_report.xlsx")
    wb = openpyxl.Workbook()
    wb.active.title = "Lookup"
    template = wb.create_sheet("Report")
    template.append(["CID", "emailaddress", "firstname", "lastname", "company"])
    wb.save(path)

    leads_df = pd.DataFrame([
        {"CID": 200, "emailaddress": "new@y.com", "firstname": "New", "lastname": "Lead", "company": "Y"},
    ])

    append_leads(path, "Report", leads_df, _field_mapping(), run_date="2026-08-08", clear_existing=True)

    wb2 = openpyxl.load_workbook(path)
    ws = wb2["Report"]
    assert ws.max_row == 2
    assert ws.cell(row=2, column=2).value == "new@y.com"


def test_append_leads_without_clear_existing_still_accumulates(tmp_path):
    # Sanity check: clear_existing defaults to False, preserving the
    # existing accumulate-by-default behavior.
    path = str(tmp_path / "lead_report.xlsx")
    wb = openpyxl.Workbook()
    wb.active.title = "Lookup"
    template = wb.create_sheet("Report")
    template.append(["CID", "emailaddress", "firstname", "lastname", "company"])
    template.append([100, "old@x.com", "Old", "Lead", "X"])
    wb.save(path)

    leads_df = pd.DataFrame([
        {"CID": 200, "emailaddress": "new@y.com", "firstname": "New", "lastname": "Lead", "company": "Y"},
    ])

    append_leads(path, "Report", leads_df, _field_mapping(), run_date="2026-08-08")

    wb2 = openpyxl.load_workbook(path)
    ws = wb2["Report"]
    assert ws.max_row == 3
    assert ws.cell(row=2, column=2).value == "old@x.com"
    assert ws.cell(row=3, column=2).value == "new@y.com"


def test_append_leads_to_template_with_reordered_and_extra_columns_appends_below_existing(tmp_path):
    # Simulates a real Lead Template: its own column order (different from the
    # leadfile's), extra leadfile columns beyond the 5 mapped fields (title,
    # industry), an existing lead already present, and that existing row's
    # formatting — all four of the client's stated requirements at once.
    path = str(tmp_path / "template.xlsx")
    wb = openpyxl.Workbook()
    wb.active.title = "Lookup"
    template = wb.create_sheet("Sheet1")
    template.append(["industry", "company", "title", "CID", "emailaddress", "firstname", "lastname"])
    existing_font = Font(bold=True, color="FF112233")
    template.append(["Retail", "Existing Co", "VP", 100, "existing@x.com", "Ex", "Isting"])
    for col in range(1, 8):
        template.cell(row=2, column=col).font = existing_font
    wb.save(path)

    leads_df = pd.DataFrame([
        {"CID": 200, "emailaddress": "new@y.com", "firstname": "New", "lastname": "Lead",
         "company": "Acme", "title": "CFO", "industry": "Finance"},
    ])

    append_leads(path, "Sheet1", leads_df, _field_mapping(), run_date="2026-08-12")

    wb2 = openpyxl.load_workbook(path)
    ws = wb2["Sheet1"]

    assert ws.max_row == 3  # appended below the existing lead, not overwriting it
    values = {ws.cell(row=1, column=c).value: ws.cell(row=3, column=c).value for c in range(1, 8)}
    assert values["industry"] == "Finance"
    assert values["company"] == "Acme"
    assert values["title"] == "CFO"
    assert values["CID"] == 200
    assert values["emailaddress"] == "new@y.com"
    assert values["firstname"] == "New"
    assert values["lastname"] == "Lead"

    # New row inherits the previous (existing) row's formatting, per column.
    for col in range(1, 8):
        assert ws.cell(row=3, column=col).font.bold is True
        assert ws.cell(row=3, column=col).font.color.rgb == "FF112233"


def test_append_leads_uses_explicit_target_field_mapping_when_header_text_differs(tmp_path):
    # A real mid-campaign Accumulated Report/Lead Template whose header text
    # doesn't match any synonym (e.g. "Email Add." instead of "emailaddress")
    # must still populate correctly once an explicit mapping is provided.
    path = str(tmp_path / "template.xlsx")
    wb = openpyxl.Workbook()
    wb.active.title = "Lookup"
    template = wb.create_sheet("Sheet1")
    template.append(["Email Add.", "Given Name", "Surname", "Org", "Campaign ID"])
    wb.save(path)

    leads_df = pd.DataFrame([
        {"CID": 200, "emailaddress": "c@y.com", "firstname": "C", "lastname": "D", "company": "Y"},
    ])
    target_mapping = FieldMapping(email="Email Add.", first_name="Given Name",
                                   last_name="Surname", company="Org", cid="Campaign ID")

    append_leads(path, "Sheet1", leads_df, _field_mapping(), run_date="2026-08-12",
                 target_field_mapping=target_mapping)

    wb2 = openpyxl.load_workbook(path)
    ws = wb2["Sheet1"]
    values = {ws.cell(row=1, column=c).value: ws.cell(row=2, column=c).value for c in range(1, 6)}
    assert values["Email Add."] == "c@y.com"
    assert values["Given Name"] == "C"
    assert values["Surname"] == "D"
    assert values["Org"] == "Y"
    assert values["Campaign ID"] == 200


def test_append_leads_without_target_mapping_falls_back_to_guess_based_matching(tmp_path):
    # No explicit mapping configured — must behave exactly as before (synonym
    # guessing / literal header match) so existing clients are unaffected.
    path = str(tmp_path / "template.xlsx")
    wb = openpyxl.Workbook()
    wb.active.title = "Lookup"
    template = wb.create_sheet("Sheet1")
    template.append(["CID", "emailaddress", "firstname", "lastname", "company"])
    wb.save(path)

    leads_df = pd.DataFrame([
        {"CID": 200, "emailaddress": "c@y.com", "firstname": "C", "lastname": "D", "company": "Y"},
    ])
    append_leads(path, "Sheet1", leads_df, _field_mapping(), run_date="2026-08-12")

    wb2 = openpyxl.load_workbook(path)
    ws = wb2["Sheet1"]
    values = {ws.cell(row=1, column=c).value: ws.cell(row=2, column=c).value for c in range(1, 6)}
    assert values["emailaddress"] == "c@y.com"
    assert values["CID"] == 200


def test_guess_target_field_mapping_matches_known_synonyms():
    guess = guess_target_field_mapping(["CID", "emailaddress", "firstname", "lastname", "company", "Extra"])
    assert guess == {"cid": "CID", "email": "emailaddress", "first_name": "firstname",
                      "last_name": "lastname", "company": "company"}


def test_guess_target_field_mapping_skips_unrecognized_headers():
    guess = guess_target_field_mapping(["Email Add.", "Given Name"])
    assert guess == {}


def test_guess_target_field_mapping_matches_snake_case_and_kebab_case():
    # Real templates commonly use "Email_Address" / "First-Name" style headers
    # instead of the literal synonym phrases.
    guess = guess_target_field_mapping(["Email_Address", "First-Name", "Last_Name", "Company_Name"])
    assert guess == {"email": "Email_Address", "first_name": "First-Name",
                      "last_name": "Last_Name", "company": "Company_Name"}


def test_find_header_row_ignores_annotation_rows_denser_than_threshold(tmp_path):
    # Regression test for a real production template: several annotation/
    # instruction rows above the real header each have a handful of scattered
    # notes (more than the old fixed 3-cell threshold), so "first row past a
    # threshold" picked one of them. Header text here deliberately matches no
    # generic synonym, so tier-1 marker matching can't short-circuit this —
    # only the density-based structural fallback can find the real row.
    path = str(tmp_path / "template.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append([])  # row 1: fully blank
    ws.append([None, "Note A", None, None, None, None, None, None, None, "Note B", None, None, None, None, "Note C", None, None, "Note D"])  # row 2: 4 scattered notes
    ws.append([None, None, "Duplicates highlighted", "Picklist", None, None, None, "Picklist", "Picklist"])  # row 3: 3 scattered notes
    ws.append(["Ref", "Email Add.", "Title", "Given Name", "Surname", "Org",
               "Role", "Dept", "Region"])  # row 4: real headers, dense (9 cells)
    wb.save(path)

    assert find_header_row(path, "Sheet1") == 4


def test_find_header_row_detects_headers_below_title_rows(tmp_path):
    path = str(tmp_path / "template.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Client Campaign Template"])
    ws.append(["Do not edit column order"])
    ws.append(["CID", "emailaddress", "firstname", "lastname", "company"])
    wb.save(path)

    assert find_header_row(path, "Sheet1") == 3
    assert read_sheet_headers(path, "Sheet1", 3) == ["CID", "emailaddress", "firstname", "lastname", "company"]


def test_find_header_row_uses_expected_headers_when_given(tmp_path):
    # A saved mapping's exact header text is a more reliable marker than the
    # generic synonym list, and is required when headers don't match any synonym.
    path = str(tmp_path / "template.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Notes"])
    ws.append(["Email Add.", "Given Name", "Campaign ID"])
    wb.save(path)

    assert find_header_row(path, "Sheet1", expected_headers=["Email Add.", "Campaign ID"]) == 2


def test_find_header_row_falls_back_to_structural_heuristic_when_no_marker_matches(tmp_path):
    # No saved mapping yet AND non-standard header text (e.g. "Email Add.")
    # that matches no generic synonym — must still find the real header row
    # by noticing it has far more non-empty cells than the title row above it.
    path = str(tmp_path / "template.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["CAMPAIGN TEMPLATE - DO NOT MODIFY LAYOUT"])
    ws.append(["Prepared by Marketing Ops"])
    ws.append(["Campaign ID", "Email Add.", "Given Name", "Surname", "Org"])
    wb.save(path)

    assert find_header_row(path, "Sheet1") == 3


def test_find_header_row_structural_fallback_handles_blank_first_column(tmp_path):
    # Regression test: when the header row's own first column (A) is blank
    # and data starts at column B, openpyxl's read-only mode represents
    # column A as an EmptyCell (no .row attribute) — the structural fallback
    # must not rely on that cell's .row.
    path = str(tmp_path / "template.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["CAMPAIGN TEMPLATE - DO NOT MODIFY LAYOUT"])
    ws.cell(row=2, column=2, value="Campaign ID")
    ws.cell(row=2, column=3, value="Email Add.")
    ws.cell(row=2, column=4, value="Given Name")
    ws.cell(row=2, column=5, value="Surname")
    wb.save(path)

    assert find_header_row(path, "Sheet1") == 2


def test_find_header_row_falls_back_to_one_when_nothing_matches(tmp_path):
    path = str(tmp_path / "template.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Foo", "Bar"])
    wb.save(path)

    assert find_header_row(path, "Sheet1") == 1


def test_append_leads_with_header_row_below_row_one_pastes_directly_under_headers(tmp_path):
    path = str(tmp_path / "template.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Client Campaign Template"])
    ws.append(["Do not edit"])
    ws.append(["CID", "emailaddress", "firstname", "lastname", "company"])
    wb.save(path)

    leads_df = pd.DataFrame([
        {"CID": 200, "emailaddress": "c@y.com", "firstname": "C", "lastname": "D", "company": "Y"},
    ])
    append_leads(path, "Sheet1", leads_df, _field_mapping(), run_date="2026-08-12", header_row=3)

    wb2 = openpyxl.load_workbook(path)
    ws2 = wb2["Sheet1"]
    assert ws2.max_row == 4  # title rows (1-2) + header (3) + one data row (4)
    values = {ws2.cell(row=3, column=c).value: ws2.cell(row=4, column=c).value for c in range(1, 6)}
    assert values["CID"] == 200
    assert values["emailaddress"] == "c@y.com"
    # rows above the header are untouched
    assert ws2.cell(row=1, column=1).value == "Client Campaign Template"
    assert ws2.cell(row=2, column=1).value == "Do not edit"


def test_append_leads_with_header_row_below_row_one_appends_below_existing_leads(tmp_path):
    path = str(tmp_path / "template.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Client Campaign Template"])
    ws.append(["CID", "emailaddress", "firstname", "lastname", "company"])
    existing_font = Font(bold=True, color="FF445566")
    ws.append([100, "existing@x.com", "Ex", "Isting", "Existing Co"])
    for col in range(1, 6):
        ws.cell(row=3, column=col).font = existing_font
    wb.save(path)

    leads_df = pd.DataFrame([
        {"CID": 200, "emailaddress": "new@y.com", "firstname": "New", "lastname": "Lead", "company": "Acme"},
    ])
    append_leads(path, "Sheet1", leads_df, _field_mapping(), run_date="2026-08-12", header_row=2)

    wb2 = openpyxl.load_workbook(path)
    ws2 = wb2["Sheet1"]
    assert ws2.max_row == 4
    values = {ws2.cell(row=2, column=c).value: ws2.cell(row=4, column=c).value for c in range(1, 6)}
    assert values["CID"] == 200
    assert values["emailaddress"] == "new@y.com"
    for col in range(1, 6):
        assert ws2.cell(row=4, column=col).font.bold is True
        assert ws2.cell(row=4, column=col).font.color.rgb == "FF445566"


def test_route_leads_by_cid_splits_into_matching_tabs_mutually_exclusively():
    leads = pd.DataFrame([
        {"CID": "119336", "emailaddress": "a@x.com"},
        {"CID": "119337", "emailaddress": "b@x.com"},
        {"CID": "119338", "emailaddress": "c@x.com"},
    ])
    tabs = [
        LeadTemplateTab(sheet_name="APAC", cids=["119336", "119337"]),
        LeadTemplateTab(sheet_name="EMEA", cids=["119338"]),
    ]

    groups, unmatched = route_leads_by_cid(leads, "CID", tabs)

    assert set(groups.keys()) == {("", "APAC"), ("", "EMEA")}
    assert list(groups[("", "APAC")]["emailaddress"]) == ["a@x.com", "b@x.com"]
    assert list(groups[("", "EMEA")]["emailaddress"]) == ["c@x.com"]
    assert unmatched.empty


def test_route_leads_by_cid_uses_tabs_own_file_path_when_set():
    leads = pd.DataFrame([
        {"CID": "119336", "emailaddress": "a@x.com"},
        {"CID": "119338", "emailaddress": "c@x.com"},
    ])
    tabs = [
        LeadTemplateTab(sheet_name="APAC", cids=["119336"]),  # falls back to default_file_path
        LeadTemplateTab(sheet_name="EMEA", cids=["119338"], file_path="C:/EMEA_only.xlsx"),
    ]

    groups, unmatched = route_leads_by_cid(leads, "CID", tabs, default_file_path="C:/shared.xlsx")

    assert set(groups.keys()) == {("C:/shared.xlsx", "APAC"), ("C:/EMEA_only.xlsx", "EMEA")}
    assert unmatched.empty


def test_route_leads_by_cid_merges_leads_routed_to_the_same_file_and_tab():
    # Two different CID groups intentionally (or accidentally) pointed at
    # the exact same destination must not silently drop one group's leads.
    leads = pd.DataFrame([
        {"CID": "1", "emailaddress": "a@x.com"},
        {"CID": "2", "emailaddress": "b@x.com"},
    ])
    tabs = [
        LeadTemplateTab(sheet_name="Shared", cids=["1"]),
        LeadTemplateTab(sheet_name="Shared", cids=["2"]),
    ]

    groups, unmatched = route_leads_by_cid(leads, "CID", tabs)

    assert set(groups.keys()) == {("", "Shared")}
    assert list(groups[("", "Shared")]["emailaddress"]) == ["a@x.com", "b@x.com"]


def test_route_leads_by_cid_returns_unmatched_leads_separately():
    leads = pd.DataFrame([
        {"CID": "119336", "emailaddress": "a@x.com"},
        {"CID": "999999", "emailaddress": "no-tab@x.com"},
    ])
    tabs = [LeadTemplateTab(sheet_name="APAC", cids=["119336"])]

    groups, unmatched = route_leads_by_cid(leads, "CID", tabs)

    assert set(groups.keys()) == {("", "APAC")}
    assert list(unmatched["emailaddress"]) == ["no-tab@x.com"]


def test_route_leads_by_cid_ignores_tabs_with_no_cids_configured():
    leads = pd.DataFrame([{"CID": "119336", "emailaddress": "a@x.com"}])
    tabs = [LeadTemplateTab(sheet_name="Catchall", cids=[])]

    groups, unmatched = route_leads_by_cid(leads, "CID", tabs)

    assert groups == {}
    assert list(unmatched["emailaddress"]) == ["a@x.com"]


def test_route_leads_by_cid_and_append_leads_write_to_correct_sheets(tmp_path):
    path = str(tmp_path / "multi_tab_template.xlsx")
    wb = openpyxl.Workbook()
    wb.active.title = "Lookup"
    apac = wb.create_sheet("APAC")
    apac.append(["CID", "emailaddress", "firstname", "lastname", "company"])
    emea = wb.create_sheet("EMEA")
    emea.append(["CID", "emailaddress", "firstname", "lastname", "company"])
    wb.save(path)

    leads = pd.DataFrame([
        {"CID": "119336", "emailaddress": "a@x.com", "firstname": "A", "lastname": "B", "company": "X"},
        {"CID": "119338", "emailaddress": "c@x.com", "firstname": "C", "lastname": "D", "company": "Y"},
    ])
    tabs = [
        LeadTemplateTab(sheet_name="APAC", cids=["119336", "119337"]),
        LeadTemplateTab(sheet_name="EMEA", cids=["119338"]),
    ]
    groups, unmatched = route_leads_by_cid(leads, "CID", tabs)
    assert unmatched.empty

    fm = _field_mapping()
    for (file_path, sheet_name), tab_leads in groups.items():
        append_leads(file_path or path, sheet_name, tab_leads, fm, run_date="2026-08-13")

    wb2 = openpyxl.load_workbook(path)
    assert wb2["APAC"].cell(row=2, column=2).value == "a@x.com"
    assert wb2["APAC"].max_row == 2
    assert wb2["EMEA"].cell(row=2, column=2).value == "c@x.com"
    assert wb2["EMEA"].max_row == 2


def test_read_sheet_headers_on_genuinely_empty_sheet_returns_empty_list_not_indexerror(tmp_path):
    # Regression test: a workbook's default/unused sheet (e.g. "Sheet" left
    # over from creating a new workbook, or any truly blank tab) has never
    # had a single cell written. openpyxl's ws[row] indexing raises
    # IndexError in that case even though max_row reports 1 — this must not
    # crash Client Setup when it defaults the sheet dropdown to such a sheet.
    path = str(tmp_path / "workbook_with_blank_sheet.xlsx")
    wb = openpyxl.Workbook()
    wb.active.title = "Blank"
    wb.save(path)

    header_row = find_header_row(path, "Blank")
    assert read_sheet_headers(path, "Blank", header_row) == []
