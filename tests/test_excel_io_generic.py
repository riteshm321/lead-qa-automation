import os
import zipfile

import openpyxl
import pandas as pd
import pytest

from core.excel_io import list_sheet_names, read_sheet_as_dataframe, backup_file, require_columns, append_leads
from core.models import FieldMapping


def _make_workbook(path: str) -> None:
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Exclusion"
    ws1.append(["Account Name", "Domain"])
    ws1.append(["Adecco UK Ltd", "adecco.co.uk"])
    ws2 = wb.create_sheet("Persona titles ")
    ws2.append(["AUDIENCE : CSUITE"])
    wb.save(path)


def test_list_sheet_names(tmp_path):
    path = str(tmp_path / "exclusion.xlsx")
    _make_workbook(path)
    assert list_sheet_names(path) == ["Exclusion", "Persona titles "]


def test_read_sheet_as_dataframe(tmp_path):
    path = str(tmp_path / "exclusion.xlsx")
    _make_workbook(path)
    df = read_sheet_as_dataframe(path, "Exclusion")
    assert list(df.columns) == ["Account Name", "Domain"]
    assert df.iloc[0]["Domain"] == "adecco.co.uk"


def test_backup_file_creates_timestamped_copy_in_backup_subfolder(tmp_path):
    path = str(tmp_path / "accumulated.xlsx")
    _make_workbook(path)

    backup_path = backup_file(path)

    assert os.path.isfile(backup_path)
    assert backup_path != path
    assert "accumulated_backup_" in os.path.basename(backup_path)
    assert os.path.basename(os.path.dirname(backup_path)) == "backup"
    assert list_sheet_names(backup_path) == list_sheet_names(path)


def test_backup_file_creates_backup_folder_when_missing(tmp_path):
    path = str(tmp_path / "accumulated.xlsx")
    _make_workbook(path)
    backup_dir = tmp_path / "backup"
    assert not backup_dir.exists()

    backup_file(path)

    assert backup_dir.is_dir()


def test_backup_file_reuses_existing_backup_folder(tmp_path):
    path = str(tmp_path / "accumulated.xlsx")
    _make_workbook(path)
    (tmp_path / "backup").mkdir()

    backup_path = backup_file(path)  # must not raise even though the folder already exists

    assert os.path.isfile(backup_path)


def test_require_columns_passes_when_all_present(tmp_path):
    path = str(tmp_path / "exclusion.xlsx")
    _make_workbook(path)
    df = read_sheet_as_dataframe(path, "Exclusion")

    require_columns(df, ["Account Name", "Domain"], file_label=path)  # should not raise


def test_require_columns_raises_clear_error_naming_file_and_column(tmp_path):
    path = str(tmp_path / "exclusion.xlsx")
    _make_workbook(path)
    df = read_sheet_as_dataframe(path, "Exclusion")

    with pytest.raises(ValueError) as exc_info:
        require_columns(df, ["Account Name", "Email"], file_label=path)

    message = str(exc_info.value)
    assert path in message
    assert "Email" in message


def test_detect_cids_from_pacing_overview_handles_offset_layout(tmp_path):
    from core.excel_io import detect_cids_from_pacing_overview

    path = str(tmp_path / "accumulated.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pacing Overview"
    # Real sheets often have their used range start at B2, not A1 — header row 3.
    ws["B2"] = "Pacing Overview"
    ws["B3"] = "SR No"
    ws["C3"] = "CID"
    ws["D3"] = "Campaign Segment"
    ws["C4"] = "118118"
    ws["D4"] = "APAC Mgr+ Q3"
    ws["C5"] = "118119"
    ws["D5"] = "EMEA Mgr+ Q3"
    ws["C6"] = "Grand Total"
    wb.save(path)

    pairs = detect_cids_from_pacing_overview(path)

    assert pairs == [("118118", "APAC Mgr+ Q3"), ("118119", "EMEA Mgr+ Q3")]


def test_detect_cids_from_pacing_overview_raises_clear_error_when_no_cid_header(tmp_path):
    from core.excel_io import detect_cids_from_pacing_overview

    path = str(tmp_path / "accumulated.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pacing Overview"
    ws["A1"] = "Nothing relevant here"
    wb.save(path)

    try:
        detect_cids_from_pacing_overview(path)
        assert False, "expected ValueError"
    except ValueError as exc:
        assert "CID" in str(exc)


def test_detect_cids_from_pacing_overview_raises_clear_error_when_sheet_missing(tmp_path):
    from core.excel_io import detect_cids_from_pacing_overview

    path = str(tmp_path / "accumulated.xlsx")
    wb = openpyxl.Workbook()
    wb.active.title = "SomeOtherSheet"
    wb.save(path)

    try:
        detect_cids_from_pacing_overview(path)
        assert False, "expected ValueError"
    except ValueError as exc:
        assert "Pacing Overview" in str(exc)


def test_detect_cids_from_pacing_overview_stops_at_grand_total_row(tmp_path):
    from core.excel_io import detect_cids_from_pacing_overview

    path = str(tmp_path / "accumulated.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pacing Overview"
    ws["B2"] = "Pacing Overview"
    ws["B3"] = "SR No"
    ws["C3"] = "CID"
    ws["D3"] = "Campaign Segment"
    ws["C4"] = "118118"
    ws["D4"] = "APAC Mgr+ Q3"
    ws["C5"] = "Grand Total"
    # A second, unrelated table below the Grand Total row must NOT be picked up.
    ws["C7"] = "999999"
    ws["D7"] = "Should Not Appear"
    wb.save(path)

    pairs = detect_cids_from_pacing_overview(path)

    assert pairs == [("118118", "APAC Mgr+ Q3")]


def test_append_leads_finds_last_row_past_bulk_formatted_empty_rows(tmp_path):
    # Real templates are often bulk-preformatted (borders/fills applied) far
    # beyond the actual data, which pushes ws.max_row way past the true last
    # lead. New leads must land right after the last row with real values,
    # not after the sheet's whole formatted range.
    path = str(tmp_path / "template.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TEMPLATE"
    ws.append(["Email_Address", "First_Name", "Last_Name", "Company_Name"])
    ws.append(["jane@x.com", "Jane", "Doe", "Acme"])
    from openpyxl.styles import PatternFill
    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for row in range(3, 500):
        ws.cell(row=row, column=1).fill = fill
    wb.save(path)
    assert openpyxl.load_workbook(path)["TEMPLATE"].max_row >= 499

    leads_df = pd.DataFrame([{"Email_Address": "bob@x.com", "First_Name": "Bob",
                               "Last_Name": "Lee", "Company_Name": "Beta"}])
    field_mapping = FieldMapping(email="Email_Address", first_name="First_Name",
                                  last_name="Last_Name", company="Company_Name", cid="")

    append_leads(path, "TEMPLATE", leads_df, field_mapping, run_date="2026-08-13")

    ws = openpyxl.load_workbook(path)["TEMPLATE"]
    assert ws.cell(row=3, column=1).value == "bob@x.com"
    assert ws.cell(row=4, column=1).value is None


def test_append_leads_matches_headers_regardless_of_separators(tmp_path):
    # Leadfile column "jobfunction" (no separator at all) must populate a
    # target header of "Job Function" (space-separated) — real leadfiles
    # frequently drop separators entirely between words.
    path = str(tmp_path / "accumulated.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Accumulated"
    ws.append(["Email_Address", "First_Name", "Last_Name", "Company_Name", "Job Function"])
    wb.save(path)

    leads_df = pd.DataFrame([{"Email_Address": "bob@x.com", "First_Name": "Bob",
                               "Last_Name": "Lee", "Company_Name": "Beta",
                               "jobfunction": "Engineering"}])
    field_mapping = FieldMapping(email="Email_Address", first_name="First_Name",
                                  last_name="Last_Name", company="Company_Name", cid="")

    append_leads(path, "Accumulated", leads_df, field_mapping, run_date="2026-08-13")

    ws = openpyxl.load_workbook(path)["Accumulated"]
    assert ws.cell(row=2, column=5).value == "Engineering"


def test_append_leads_matches_headers_with_extra_suffix_via_containment(tmp_path):
    # "MarketSegmentReferential" (leadfile) must populate "Market Segment"
    # (target) — export tools often append noise like "Referential".
    path = str(tmp_path / "accumulated.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Accumulated"
    ws.append(["Email_Address", "First_Name", "Last_Name", "Company_Name", "Market Segment", "I am a"])
    wb.save(path)

    leads_df = pd.DataFrame([{"Email_Address": "bob@x.com", "First_Name": "Bob",
                               "Last_Name": "Lee", "Company_Name": "Beta",
                               "MarketSegmentReferential": "Enterprise",
                               "IAMAReferential": "Decision Maker"}])
    field_mapping = FieldMapping(email="Email_Address", first_name="First_Name",
                                  last_name="Last_Name", company="Company_Name", cid="")

    unmatched = append_leads(path, "Accumulated", leads_df, field_mapping, run_date="2026-08-13")

    ws = openpyxl.load_workbook(path)["Accumulated"]
    assert ws.cell(row=2, column=5).value == "Enterprise"
    assert ws.cell(row=2, column=6).value == "Decision Maker"
    assert unmatched == []


def test_append_leads_matches_headers_via_known_synonym_group(tmp_path):
    # Regression test: leadfile "companysize" and Lead Template "Employee
    # Size" are genuinely different words for the same field — no amount of
    # string-similarity scoring matches them ("company" vs "employee" score
    # well under the fuzzy threshold), so this previously left the column
    # blank despite both files having equivalent data.
    path = str(tmp_path / "accumulated.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Accumulated"
    ws.append(["Email_Address", "First_Name", "Last_Name", "Company_Name", "Employee Size"])
    wb.save(path)

    leads_df = pd.DataFrame([{"Email_Address": "bob@x.com", "First_Name": "Bob",
                               "Last_Name": "Lee", "Company_Name": "Beta",
                               "companysize": "1,000-4,999"}])
    field_mapping = FieldMapping(email="Email_Address", first_name="First_Name",
                                  last_name="Last_Name", company="Company_Name", cid="")

    unmatched = append_leads(path, "Accumulated", leads_df, field_mapping, run_date="2026-08-13")

    ws = openpyxl.load_workbook(path)["Accumulated"]
    assert ws.cell(row=2, column=5).value == "1,000-4,999"
    assert unmatched == []


def test_append_leads_leaves_ambiguous_containment_matches_unmatched(tmp_path):
    # Two leadfile columns both contain "region" — auto-picking either one
    # risks silently wiring the wrong data into a client's real report, so
    # neither should be auto-matched.
    path = str(tmp_path / "accumulated.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Accumulated"
    ws.append(["Email_Address", "First_Name", "Last_Name", "Company_Name", "Region"])
    wb.save(path)

    leads_df = pd.DataFrame([{"Email_Address": "bob@x.com", "First_Name": "Bob",
                               "Last_Name": "Lee", "Company_Name": "Beta",
                               "SalesRegion": "APAC", "ShippingRegion": "EMEA"}])
    field_mapping = FieldMapping(email="Email_Address", first_name="First_Name",
                                  last_name="Last_Name", company="Company_Name", cid="")

    unmatched = append_leads(path, "Accumulated", leads_df, field_mapping, run_date="2026-08-13")

    ws = openpyxl.load_workbook(path)["Accumulated"]
    assert ws.cell(row=2, column=5).value is None
    assert unmatched == ["Region"]


def test_append_leads_preserves_external_link_parts_byte_for_byte(tmp_path):
    path = str(tmp_path / "template.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TEMPLATE"
    ws.append(["Email_Address", "First_Name", "Last_Name", "Company_Name"])
    wb.save(path)

    fake_external_link = b"<not real xml, just needs to round-trip untouched>"
    with zipfile.ZipFile(path, "a", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("xl/externalLinks/externalLink1.xml", fake_external_link)

    leads_df = pd.DataFrame([{"Email_Address": "bob@x.com", "First_Name": "Bob",
                               "Last_Name": "Lee", "Company_Name": "Beta"}])
    field_mapping = FieldMapping(email="Email_Address", first_name="First_Name",
                                  last_name="Last_Name", company="Company_Name", cid="")

    append_leads(path, "TEMPLATE", leads_df, field_mapping, run_date="2026-08-13")

    with zipfile.ZipFile(path, "r") as zf:
        assert zf.read("xl/externalLinks/externalLink1.xml") == fake_external_link


def test_detect_cids_from_pacing_overview_stops_at_first_blank_cid_row(tmp_path):
    from core.excel_io import detect_cids_from_pacing_overview

    path = str(tmp_path / "accumulated.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pacing Overview"
    ws["B2"] = "Pacing Overview"
    ws["B3"] = "SR No"
    ws["C3"] = "CID"
    ws["D3"] = "Campaign Segment"
    ws["C4"] = "118118"
    ws["D4"] = "APAC Mgr+ Q3"
    # Row 5 has a blank CID cell — scanning must stop here.
    ws["D5"] = "Blank CID row"
    # A later, unrelated CID-labeled row must NOT be picked up.
    ws["C6"] = "999999"
    ws["D6"] = "Should Not Appear"
    wb.save(path)

    pairs = detect_cids_from_pacing_overview(path)

    assert pairs == [("118118", "APAC Mgr+ Q3")]


def test_read_pacing_overview_table_captures_every_column(tmp_path):
    from core.excel_io import read_pacing_overview_table

    path = str(tmp_path / "accumulated.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pacing Overview"
    ws["B2"] = "Pacing Overview"
    ws["B3"] = "SR No"
    ws["C3"] = "CID"
    ws["D3"] = "Campaign Segment"
    ws["E3"] = "Target"
    ws["F3"] = "Delivered"
    ws["B4"] = 1
    ws["C4"] = "118118"
    ws["D4"] = "APAC Mgr+ Q3"
    ws["E4"] = 100
    ws["F4"] = 42
    ws["B5"] = 2
    ws["C5"] = "118119"
    ws["D5"] = "EMEA Mgr+ Q3"
    ws["E5"] = 80
    ws["F5"] = 80
    ws["C6"] = "Grand Total"
    ws["E6"] = 180
    ws["F6"] = 122
    wb.save(path)

    df = read_pacing_overview_table(path)

    assert list(df.columns) == ["SR No", "CID", "Campaign Segment", "Target", "Delivered"]
    assert df.iloc[0].tolist() == [1, "118118", "APAC Mgr+ Q3", 100, 42]
    assert df.iloc[1].tolist() == [2, "118119", "EMEA Mgr+ Q3", 80, 80]
    # The Grand Total row is kept as a real summary row, not dropped.
    assert len(df) == 3
    assert df.iloc[2].tolist() == ["", "Grand Total", "", 180, 122]


def test_read_pacing_overview_table_formats_pacing_column_as_percentage(tmp_path):
    from core.excel_io import read_pacing_overview_table

    path = str(tmp_path / "accumulated.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pacing Overview"
    ws["B3"] = "CID"
    ws["C3"] = "Pacing"
    ws["B4"] = "118118"
    ws["C4"] = 0.0547368421
    wb.save(path)

    df = read_pacing_overview_table(path)

    assert df.iloc[0]["Pacing"] == "5%"


def test_read_pacing_overview_table_formats_date_headers_without_timestamp(tmp_path):
    import datetime as dt

    from core.excel_io import read_pacing_overview_table

    path = str(tmp_path / "accumulated.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pacing Overview"
    ws["B3"] = "CID"
    ws["C3"] = dt.datetime(2026, 8, 19)
    ws["B4"] = "118118"
    ws["C4"] = 4
    wb.save(path)

    df = read_pacing_overview_table(path)

    assert list(df.columns) == ["CID", "19-Aug"]


def test_read_pacing_overview_table_skips_hidden_date_columns(tmp_path):
    # Regression test: a date column the sheet itself has hidden (e.g. an
    # old date the client collapsed to reduce clutter) must not appear in
    # the Jira table — openpyxl reads a hidden column's values the same as
    # any visible one, so without an explicit check it would show a date
    # the sheet doesn't visibly have.
    import datetime as dt

    from core.excel_io import read_pacing_overview_table

    path = str(tmp_path / "accumulated.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pacing Overview"
    ws["B3"] = "CID"
    ws["C3"] = dt.datetime(2026, 8, 18)
    ws["D3"] = dt.datetime(2026, 8, 19)
    ws["B4"] = "118118"
    ws["C4"] = 3
    ws["D4"] = 4
    ws.column_dimensions["C"].hidden = True
    wb.save(path)

    df = read_pacing_overview_table(path)

    assert list(df.columns) == ["CID", "19-Aug"]
    assert df.iloc[0].tolist() == ["118118", 4]


def test_read_pacing_overview_table_stops_at_blank_cid_row(tmp_path):
    from core.excel_io import read_pacing_overview_table

    path = str(tmp_path / "accumulated.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pacing Overview"
    ws["C3"] = "CID"
    ws["D3"] = "Campaign Segment"
    ws["C4"] = "118118"
    ws["D4"] = "APAC Mgr+ Q3"
    ws["D5"] = "Blank CID row"
    ws["C6"] = "999999"
    ws["D6"] = "Should Not Appear"
    wb.save(path)

    df = read_pacing_overview_table(path)

    assert len(df) == 1
    assert df.iloc[0]["CID"] == "118118"


def test_read_pacing_overview_table_raises_clear_error_when_sheet_missing(tmp_path):
    from core.excel_io import read_pacing_overview_table

    path = str(tmp_path / "accumulated.xlsx")
    wb = openpyxl.Workbook()
    wb.active.title = "SomeOtherSheet"
    wb.save(path)

    with pytest.raises(ValueError, match="Pacing Overview"):
        read_pacing_overview_table(path)
