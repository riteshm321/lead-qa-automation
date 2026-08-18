import io

import openpyxl
import pandas as pd

from core.excel_io import read_leadfile


def _upload(name: str, content: bytes) -> io.BytesIO:
    f = io.BytesIO(content)
    f.name = name
    return f


def test_read_leadfile_plain_utf8_csv():
    df = read_leadfile(_upload("leads.csv", b"Email,First\nabc@x.com,Bob\n"))
    assert list(df.columns) == ["Email", "First"]
    assert df.iloc[0]["Email"] == "abc@x.com"


def test_read_leadfile_strips_utf8_bom():
    content = "Email,First\nabc@x.com,Bob\n".encode("utf-8-sig")
    df = read_leadfile(_upload("leads.csv", content))
    # A leaked BOM would show up as "﻿Email" instead of "Email".
    assert list(df.columns) == ["Email", "First"]


def test_read_leadfile_detects_semicolon_delimiter():
    df = read_leadfile(_upload("leads.csv", b"Email;First\nabc@x.com;Bob\n"))
    assert list(df.columns) == ["Email", "First"]
    assert df.iloc[0]["First"] == "Bob"


def test_read_leadfile_detects_tab_delimiter():
    df = read_leadfile(_upload("leads.csv", b"Email\tFirst\nabc@x.com\tBob\n"))
    assert list(df.columns) == ["Email", "First"]


def test_read_leadfile_falls_back_to_cp1252_for_special_characters():
    content = "Email,Company\nabc@x.com,Café Corp\n".encode("cp1252")
    df = read_leadfile(_upload("leads.csv", content))
    assert df.iloc[0]["Company"] == "Café Corp"


def test_read_leadfile_xlsx_still_works(tmp_path):
    path = tmp_path / "leads.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Email", "First"])
    ws.append(["abc@x.com", "Bob"])
    wb.save(path)

    with open(path, "rb") as fh:
        df = read_leadfile(_upload("leads.xlsx", fh.read()))

    assert list(df.columns) == ["Email", "First"]
    assert df.iloc[0]["Email"] == "abc@x.com"


def test_read_leadfile_xlsx_uses_active_sheet_not_first_by_position(tmp_path):
    # Regression test: a real client file had a pivot-summary sheet
    # ("Sheet2") positioned before the real leads sheet ("Sheet1"), with
    # "Sheet1" set as the active sheet. pd.read_excel's default
    # sheet_name=0 (and openpyxl's sheetnames[0]) picked the summary sheet
    # by position instead, whose row 1 is blank -> "Unnamed: 0"/"Unnamed: 1"
    # column names, even though the real header row visibly looked fine to
    # the user when they opened the file in Excel (which shows the active
    # sheet, "Sheet1").
    path = tmp_path / "leads.xlsx"
    wb = openpyxl.Workbook()
    summary_ws = wb.active
    summary_ws.title = "Sheet2"
    summary_ws.append([None, None])
    summary_ws.append(["Row Labels", "Count of emailAddress"])

    leads_ws = wb.create_sheet("Sheet1")
    leads_ws.append(["Email", "First Name"])
    leads_ws.append(["abc@x.com", "Bob"])
    wb.active = wb.sheetnames.index("Sheet1")
    wb.save(path)

    with open(path, "rb") as fh:
        df = read_leadfile(_upload("leads.xlsx", fh.read()))

    assert list(df.columns) == ["Email", "First Name"]
    assert df.iloc[0]["Email"] == "abc@x.com"


def test_read_leadfile_xlsx_skips_title_row_above_real_header(tmp_path):
    # Regression test: a New Leads export with a title/instructions row
    # above the real header previously got read with header=0, so pandas
    # generated "Unnamed: 0"/"Unnamed: 1" column names from that blank title
    # row instead of the real "Email"/"First Name"/... headers below it.
    path = tmp_path / "leads.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Lead Report - Generated 2026-08-18"])
    ws.append(["Email", "First Name", "Last Name", "Company", "CID"])
    ws.append(["abc@x.com", "Bob", "Smith", "Acme", "12345"])
    wb.save(path)

    with open(path, "rb") as fh:
        df = read_leadfile(_upload("leads.xlsx", fh.read()))

    assert list(df.columns) == ["Email", "First Name", "Last Name", "Company", "CID"]
    assert df.iloc[0]["Email"] == "abc@x.com"
