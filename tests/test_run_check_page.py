import datetime
import os
from unittest.mock import patch

import openpyxl
import pandas as pd
from streamlit.testing.v1 import AppTest

from core.app_settings import get_clients_dir, save_jira_settings
from core.jira_client import JiraError
from core.models import ClientProfile, FieldMapping, DuplicateConfig, LeadTemplateTab, ComplexAccountConfig
from core.pipeline import PipelineResult, run_pipeline
from core.profile_store import save_profile

_PAGE_PATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), "pages", "2_Run_Check.py")


def _make_accumulated_report(path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Accumulated"
    ws.append(["Email_Address", "First_Name", "Last_Name", "Company_Name", "CID", "Date"])
    ws.append(["existing@dup.com", "Existing", "Person", "DupCo", "1", "2026-08-01"])
    wb.create_sheet("Refund").append(["Email_Address", "First_Name", "Last_Name", "Company_Name", "CID", "Date"])
    wb.save(path)


def test_cached_loaders_hash_their_mtime_argument():
    # Regression test for a real, confirmed bug: Streamlit's @st.cache_data
    # silently EXCLUDES any parameter whose name starts with an underscore
    # from the cache key hash. _cached_tal_index/_cached_asset_specs pass
    # os.path.getmtime(path) specifically to bust the cache when the
    # underlying file changes mid-session — naming that parameter "_mtime"
    # made Streamlit ignore it entirely, so a file edited and re-saved to
    # the same path during a running session kept silently serving the
    # first-loaded (now-stale) data. Verified by parsing the actual page
    # source (can't import a page module whose filename starts with a
    # digit) rather than re-deriving the bug in an unrelated toy function.
    import ast

    with open(_PAGE_PATH, encoding="utf-8") as f:
        tree = ast.parse(f.read(), filename=_PAGE_PATH)

    checked_any = False
    for node in ast.walk(tree):
        if not isinstance(node, ast.FunctionDef):
            continue
        is_cached = any(
            (isinstance(d, ast.Call) and getattr(d.func, "attr", None) == "cache_data")
            or (isinstance(d, ast.Attribute) and d.attr == "cache_data")
            for d in node.decorator_list
        )
        if not is_cached:
            continue
        checked_any = True
        param_names = [a.arg for a in node.args.args]
        underscored = [p for p in param_names if p.startswith("_")]
        assert not underscored, (
            f"@st.cache_data function '{node.name}' has underscore-prefixed param(s) "
            f"{underscored} — Streamlit excludes these from the cache key, silently "
            f"breaking any cache-busting argument (e.g. a file's mtime) passed there."
        )

    assert checked_any, "expected at least one @st.cache_data-decorated function in this page"


def test_approved_refund_lead_lands_in_accumulated_tab_not_just_refund(tmp_path, monkeypatch):
    # End-to-end regression test for the "approve a refunded lead as valid"
    # feature: AppTest can't simulate a real file upload, so this pre-seeds
    # session_state exactly as it looks right after a real Run Check click
    # (one valid lead, one refund-flagged lead), then drives the actual
    # Refund Reasons checkbox and Finalize button.
    monkeypatch.chdir(tmp_path)
    acc_path = str(tmp_path / "accumulated.xlsx")
    _make_accumulated_report(acc_path)

    fm = FieldMapping(email="Email_Address", first_name="First_Name", last_name="Last_Name",
                       company="Company_Name", cid="CID")
    profile = ClientProfile(
        name="Test Client",
        accumulated_report_path=acc_path,
        field_mapping=fm,
        duplicate=DuplicateConfig(enabled=True),
    )
    save_profile(profile, get_clients_dir())

    new_leads = pd.DataFrame([
        {"Email_Address": "bob@new.com", "First_Name": "Bob", "Last_Name": "Lee", "Company_Name": "Beta", "CID": "1"},
        {"Email_Address": "existing@dup.com", "First_Name": "Existing", "Last_Name": "Person",
         "Company_Name": "DupCo", "CID": "1"},
    ])
    result = PipelineResult(valid_indices=[0], refund_reasons={1: "Duplicate - exact email"})

    at = AppTest.from_file(_PAGE_PATH, default_timeout=15)
    at.session_state["run_new_leads"] = new_leads
    at.session_state["run_result"] = result
    at.session_state["run_result_for"] = "Test Client"
    at.run()
    assert not at.exception

    # AppTest has no widget accessor for st.data_editor (and Streamlit
    # forbids writing a data_editor's own key via session_state directly),
    # so simulate "the user ticked the one refund row's checkbox" by
    # patching st.data_editor itself to return the edited table the real
    # widget would have, for this one render.
    edited_table = pd.DataFrame([{
        "Approve as valid": True, "Row": 3, "Email": "existing@dup.com",
        "Company": "DupCo", "CID": "1", "Reason": "Duplicate - exact email",
    }])
    with patch("streamlit.data_editor", return_value=edited_table):
        finalize_button = next(b for b in at.button if b.label == "Finalize")
        finalize_button.click().run()
    assert not at.exception

    wb = openpyxl.load_workbook(acc_path)
    acc_rows = [tuple(r) for r in wb["Accumulated"].iter_rows(min_row=2, values_only=True) if r[0] is not None]
    refund_rows = [tuple(r) for r in wb["Refund"].iter_rows(min_row=2, values_only=True) if r[0] is not None]

    acc_emails = {row[0] for row in acc_rows}
    refund_emails = {row[0] for row in refund_rows}

    assert "bob@new.com" in acc_emails
    assert "existing@dup.com" in acc_emails  # approved despite being auto-flagged for refund
    assert refund_emails == set()  # nothing left in Refund tab once approved


def test_unapproved_refund_lead_stays_refund_only(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    acc_path = str(tmp_path / "accumulated.xlsx")
    _make_accumulated_report(acc_path)

    fm = FieldMapping(email="Email_Address", first_name="First_Name", last_name="Last_Name",
                       company="Company_Name", cid="CID")
    profile = ClientProfile(
        name="Test Client",
        accumulated_report_path=acc_path,
        field_mapping=fm,
        duplicate=DuplicateConfig(enabled=True),
    )
    save_profile(profile, get_clients_dir())

    new_leads = pd.DataFrame([
        {"Email_Address": "existing@dup.com", "First_Name": "Existing", "Last_Name": "Person",
         "Company_Name": "DupCo", "CID": "1"},
    ])
    result = PipelineResult(valid_indices=[], refund_reasons={0: "Duplicate - exact email"})

    at = AppTest.from_file(_PAGE_PATH, default_timeout=15)
    at.session_state["run_new_leads"] = new_leads
    at.session_state["run_result"] = result
    at.session_state["run_result_for"] = "Test Client"
    at.run()

    # Leave the checkbox unticked and finalize directly.
    finalize_button = next(b for b in at.button if b.label == "Finalize")
    finalize_button.click().run()
    assert not at.exception

    wb = openpyxl.load_workbook(acc_path)
    acc_rows = [r for r in wb["Accumulated"].iter_rows(min_row=2, values_only=True) if r[0] is not None]
    refund_rows = [r for r in wb["Refund"].iter_rows(min_row=2, values_only=True) if r[0] is not None]

    assert len(acc_rows) == 1  # only the pre-existing accumulated row, nothing new added
    assert len(refund_rows) == 1
    assert refund_rows[0][0] == "existing@dup.com"


def test_select_all_as_valid_approves_every_refund_lead(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    acc_path = str(tmp_path / "accumulated.xlsx")
    _make_accumulated_report(acc_path)

    fm = FieldMapping(email="Email_Address", first_name="First_Name", last_name="Last_Name",
                       company="Company_Name", cid="CID")
    profile = ClientProfile(
        name="Test Client",
        accumulated_report_path=acc_path,
        field_mapping=fm,
        duplicate=DuplicateConfig(enabled=True),
    )
    save_profile(profile, get_clients_dir())

    new_leads = pd.DataFrame([
        {"Email_Address": "a@x.com", "First_Name": "A", "Last_Name": "One", "Company_Name": "X", "CID": "1"},
        {"Email_Address": "b@x.com", "First_Name": "B", "Last_Name": "Two", "Company_Name": "X", "CID": "1"},
    ])
    result = PipelineResult(valid_indices=[], refund_reasons={0: "Exclusion - domain", 1: "Exclusion - domain"})

    at = AppTest.from_file(_PAGE_PATH, default_timeout=15)
    at.session_state["run_new_leads"] = new_leads
    at.session_state["run_result"] = result
    at.session_state["run_result_for"] = "Test Client"
    at.run()

    select_all = next(b for b in at.button if b.label == "Select all as valid")
    select_all.click().run()
    assert not at.exception

    finalize_button = next(b for b in at.button if b.label == "Finalize")
    finalize_button.click().run()
    assert not at.exception

    wb = openpyxl.load_workbook(acc_path)
    acc_emails = {r[0] for r in wb["Accumulated"].iter_rows(min_row=2, values_only=True) if r[0] is not None}
    refund_rows = [r for r in wb["Refund"].iter_rows(min_row=2, values_only=True) if r[0] is not None]

    assert {"a@x.com", "b@x.com"} <= acc_emails
    assert refund_rows == []


def test_post_summary_to_jira_after_finalize(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    acc_path = str(tmp_path / "accumulated.xlsx")
    _make_accumulated_report(acc_path)
    save_jira_settings("https://example.atlassian.net", "me@example.com", "token123")

    fm = FieldMapping(email="Email_Address", first_name="First_Name", last_name="Last_Name",
                       company="Company_Name", cid="CID")
    profile = ClientProfile(
        name="Test Client",
        accumulated_report_path=acc_path,
        field_mapping=fm,
        duplicate=DuplicateConfig(enabled=True),
        jira_ticket_key="PROJ-1234",
    )
    save_profile(profile, get_clients_dir())

    new_leads = pd.DataFrame([
        {"Email_Address": "bob@new.com", "First_Name": "Bob", "Last_Name": "Lee", "Company_Name": "Beta", "CID": "1"},
    ])
    result = PipelineResult(valid_indices=[0], refund_reasons={})

    at = AppTest.from_file(_PAGE_PATH, default_timeout=15)
    at.session_state["run_new_leads"] = new_leads
    at.session_state["run_result"] = result
    at.session_state["run_result_for"] = "Test Client"
    at.run()

    finalize_button = next(b for b in at.button if b.label == "Finalize")
    finalize_button.click().run()
    assert not at.exception

    # The "Post to Jira" prompt should now be showing, pre-filled with a
    # summary — verify it before actually posting anything.
    import datetime as _dt
    opening_box = next(t for t in at.text_area if t.key == "jira_comment_opening")
    assert _dt.date.today().strftime("%d-%m-%y") in opening_box.value
    assert "PFB summary for the Lead QA dated" in opening_box.value
    closing_box = next(t for t in at.text_area if t.key == "jira_comment_closing")
    assert closing_box.value == "Thanks"

    # The Accumulated File link checkbox should be offered (ticked by default).
    link_checkbox = next(c for c in at.checkbox if c.key == "jira_link_Accumulated File")
    assert link_checkbox.value is True

    with patch("core.jira_client.post_comment_body") as mock_post:
        post_button = next(b for b in at.button if b.key == "jira_post_button")
        post_button.click().run()
        assert not at.exception

    mock_post.assert_called_once()
    call_args = mock_post.call_args[0]
    assert call_args[0] == "https://example.atlassian.net"
    assert call_args[1] == "me@example.com"
    assert call_args[2] == "token123"
    assert call_args[3] == "PROJ-1234"
    adf_body = call_args[4]
    assert adf_body["type"] == "doc"
    # The file link should show up as a real ADF link mark, not plain text.
    all_marks = [
        mark
        for node in adf_body["content"] if node["type"] == "orderedList"
        for item in node["content"]
        for para in item["content"]
        for text_node in para["content"]
        for mark in text_node.get("marks", [])
    ]
    assert any(mark["type"] == "link" for mark in all_marks)

    # Session state cleaned up so the prompt disappears after a successful post.
    assert "last_finalized_summary" not in at.session_state


def test_jira_post_uploads_provided_attachment(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    acc_path = str(tmp_path / "accumulated.xlsx")
    _make_accumulated_report(acc_path)
    save_jira_settings("https://example.atlassian.net", "me@example.com", "token123")

    fm = FieldMapping(email="Email_Address", first_name="First_Name", last_name="Last_Name",
                       company="Company_Name", cid="CID")
    profile = ClientProfile(
        name="Test Client", accumulated_report_path=acc_path, field_mapping=fm,
        duplicate=DuplicateConfig(enabled=True), jira_ticket_key="PROJ-1234",
    )
    save_profile(profile, get_clients_dir())

    new_leads = pd.DataFrame([
        {"Email_Address": "bob@new.com", "First_Name": "Bob", "Last_Name": "Lee", "Company_Name": "Beta", "CID": "1"},
    ])
    result = PipelineResult(valid_indices=[0], refund_reasons={})

    at = AppTest.from_file(_PAGE_PATH, default_timeout=15)
    at.session_state["run_new_leads"] = new_leads
    at.session_state["run_result"] = result
    at.session_state["run_result_for"] = "Test Client"
    at.run()

    finalize_button = next(b for b in at.button if b.label == "Finalize")
    finalize_button.click().run()
    assert not at.exception

    # Simulate a file already selected via st.file_uploader (AppTest can't
    # drive a real file upload — see the established limitation noted
    # elsewhere in this suite) by pre-seeding what the uploader branch
    # writes into session_state.
    at.session_state["jira_attachment_bytes"] = b"fake-file-bytes"
    at.session_state["jira_attachment_name"] = "notes.txt"

    with patch("core.jira_client.post_comment_body") as mock_post_comment, \
         patch("core.jira_client.upload_attachment") as mock_upload:
        post_button = next(b for b in at.button if b.key == "jira_post_button")
        post_button.click().run()
        assert not at.exception

    mock_post_comment.assert_called_once()
    mock_upload.assert_called_once()
    assert mock_upload.call_args[0][4] == "notes.txt"
    assert mock_upload.call_args[0][5] == b"fake-file-bytes"
    assert "jira_attachment_bytes" not in at.session_state


def test_jira_post_reports_attachment_failure_without_blocking_comment(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    acc_path = str(tmp_path / "accumulated.xlsx")
    _make_accumulated_report(acc_path)
    save_jira_settings("https://example.atlassian.net", "me@example.com", "token123")

    fm = FieldMapping(email="Email_Address", first_name="First_Name", last_name="Last_Name",
                       company="Company_Name", cid="CID")
    profile = ClientProfile(
        name="Test Client", accumulated_report_path=acc_path, field_mapping=fm,
        duplicate=DuplicateConfig(enabled=True), jira_ticket_key="PROJ-1234",
    )
    save_profile(profile, get_clients_dir())

    new_leads = pd.DataFrame([
        {"Email_Address": "bob@new.com", "First_Name": "Bob", "Last_Name": "Lee", "Company_Name": "Beta", "CID": "1"},
    ])
    result = PipelineResult(valid_indices=[0], refund_reasons={})

    # This test drives more sequential button-click+rerun cycles than any
    # other in this file (Finalize, post, retry), which occasionally brushes
    # up against AppTest's default wait window under momentary system load —
    # a longer timeout here is slack for the test harness, not a change to
    # the app's own behavior.
    at = AppTest.from_file(_PAGE_PATH, default_timeout=30)
    at.session_state["run_new_leads"] = new_leads
    at.session_state["run_result"] = result
    at.session_state["run_result_for"] = "Test Client"
    at.run()

    finalize_button = next(b for b in at.button if b.label == "Finalize")
    finalize_button.click().run()
    assert not at.exception

    at.session_state["jira_attachment_bytes"] = b"fake-file-bytes"
    at.session_state["jira_attachment_name"] = "notes.txt"

    with patch("core.jira_client.post_comment_body") as mock_post_comment, \
         patch("core.jira_client.upload_attachment", side_effect=JiraError("boom")):
        post_button = next(b for b in at.button if b.key == "jira_post_button")
        post_button.click().run()
        assert not at.exception

    mock_post_comment.assert_called_once()
    assert any("boom" in e.value for e in at.error)
    # The comment succeeded, so its success message must still show, and
    # last_finalized_summary must stay put so the retry button can appear.
    assert any("Posted to PROJ-1234" in s.value for s in at.success)
    assert "last_finalized_summary" in at.session_state

    # Retrying must not repost the comment — only the failed attachment.
    with patch("core.jira_client.post_comment_body") as mock_post_comment_retry, \
         patch("core.jira_client.upload_attachment") as mock_upload_retry:
        retry_button = next(b for b in at.button if b.key == "jira_post_button")
        assert "Retry failed attachment" in retry_button.label
        retry_button.click().run()
        assert not at.exception

    mock_post_comment_retry.assert_not_called()
    mock_upload_retry.assert_called_once()
    assert "last_finalized_summary" not in at.session_state


def test_jira_prompt_does_not_appear_without_ticket_key(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    acc_path = str(tmp_path / "accumulated.xlsx")
    _make_accumulated_report(acc_path)

    fm = FieldMapping(email="Email_Address", first_name="First_Name", last_name="Last_Name",
                       company="Company_Name", cid="CID")
    profile = ClientProfile(
        name="Test Client",
        accumulated_report_path=acc_path,
        field_mapping=fm,
        duplicate=DuplicateConfig(enabled=True),
    )
    save_profile(profile, get_clients_dir())

    new_leads = pd.DataFrame([
        {"Email_Address": "bob@new.com", "First_Name": "Bob", "Last_Name": "Lee", "Company_Name": "Beta", "CID": "1"},
    ])
    result = PipelineResult(valid_indices=[0], refund_reasons={})

    at = AppTest.from_file(_PAGE_PATH, default_timeout=15)
    at.session_state["run_new_leads"] = new_leads
    at.session_state["run_result"] = result
    at.session_state["run_result_for"] = "Test Client"
    at.run()

    finalize_button = next(b for b in at.button if b.label == "Finalize")
    finalize_button.click().run()
    assert not at.exception

    assert not any(t.key == "jira_comment_opening" for t in at.text_area)


def test_post_summary_to_jira_includes_pacing_overview_as_native_table(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    acc_path = str(tmp_path / "accumulated.xlsx")
    _make_accumulated_report(acc_path)
    save_jira_settings("https://example.atlassian.net", "me@example.com", "token123")

    wb = openpyxl.load_workbook(acc_path)
    pacing = wb.create_sheet("Pacing Overview")
    pacing.append(["SR No", "CID", "Campaign Segment"])
    pacing.append([1, "118118", "APAC Mgr+ Q3"])
    pacing.append([2, "118119", "EMEA Mgr+ Q3"])
    wb.save(acc_path)

    fm = FieldMapping(email="Email_Address", first_name="First_Name", last_name="Last_Name",
                       company="Company_Name", cid="CID")
    profile = ClientProfile(
        name="Test Client",
        accumulated_report_path=acc_path,
        field_mapping=fm,
        duplicate=DuplicateConfig(enabled=True),
        jira_ticket_key="PROJ-1234",
        jira_reporter_name="Jane",
    )
    save_profile(profile, get_clients_dir())

    new_leads = pd.DataFrame([
        {"Email_Address": "bob@new.com", "First_Name": "Bob", "Last_Name": "Lee", "Company_Name": "Beta", "CID": "1"},
    ])
    result = PipelineResult(valid_indices=[0], refund_reasons={})

    at = AppTest.from_file(_PAGE_PATH, default_timeout=15)
    at.session_state["run_new_leads"] = new_leads
    at.session_state["run_result"] = result
    at.session_state["run_result_for"] = "Test Client"
    at.run()

    finalize_button = next(b for b in at.button if b.label == "Finalize")
    finalize_button.click().run()
    assert not at.exception

    opening_box = next(t for t in at.text_area if t.key == "jira_comment_opening")
    assert "Hi Jane" in opening_box.value

    pacing_checkbox = next(c for c in at.checkbox if c.key == "jira_include_pacing")
    assert pacing_checkbox.value is True

    with patch("core.jira_client.post_comment_body") as mock_post:
        post_button = next(b for b in at.button if b.key == "jira_post_button")
        post_button.click().run()
        assert not at.exception

    adf_body = mock_post.call_args[0][4]
    table_nodes = [n for n in adf_body["content"] if n["type"] == "table"]
    assert len(table_nodes) == 1
    header_row = table_nodes[0]["content"][0]
    header_texts = [c["content"][0]["content"][0]["text"] for c in header_row["content"]]
    assert header_texts == ["SR No", "CID", "Campaign Segment"]
    data_row_1 = table_nodes[0]["content"][1]
    assert data_row_1["content"][1]["content"][0]["content"][0]["text"] == "118118"


def test_jira_summary_includes_lead_report_link_for_lead_qa_mode_with_template(tmp_path, monkeypatch):
    # Regression test: the "Lead Report" file link must key off client_mode
    # == "Lead QA" (the mode that actually has lead_template_path set —
    # counterintuitively, "Lead QA & Upload" mode has no Lead Template at
    # all), not "Lead QA & Upload".
    monkeypatch.chdir(tmp_path)
    acc_path = str(tmp_path / "accumulated.xlsx")
    _make_accumulated_report(acc_path)
    template_path = str(tmp_path / "lead_report.xlsx")
    wb = openpyxl.Workbook()
    wb.active.append(["Email_Address", "First_Name", "Last_Name", "Company_Name", "CID"])
    wb.save(template_path)
    save_jira_settings("https://example.atlassian.net", "me@example.com", "token123")

    fm = FieldMapping(email="Email_Address", first_name="First_Name", last_name="Last_Name",
                       company="Company_Name", cid="CID")
    profile = ClientProfile(
        name="Test Client",
        accumulated_report_path=acc_path,
        field_mapping=fm,
        duplicate=DuplicateConfig(enabled=True),
        jira_ticket_key="PROJ-1234",
        client_mode="Lead QA",
        lead_template_path=template_path,
        lead_template_sheet_name="Sheet",
    )
    save_profile(profile, get_clients_dir())

    new_leads = pd.DataFrame([
        {"Email_Address": "bob@new.com", "First_Name": "Bob", "Last_Name": "Lee", "Company_Name": "Beta", "CID": "1"},
    ])
    result = PipelineResult(valid_indices=[0], refund_reasons={})

    at = AppTest.from_file(_PAGE_PATH, default_timeout=15)
    at.session_state["run_new_leads"] = new_leads
    at.session_state["run_result"] = result
    at.session_state["run_result_for"] = "Test Client"
    at.run()

    finalize_button = next(b for b in at.button if b.label == "Finalize")
    finalize_button.click().run()
    assert not at.exception

    assert any(c.key == "jira_link_Lead Report" for c in at.checkbox)


def test_jira_summary_omits_lead_report_link_for_lead_qa_and_upload_mode(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    acc_path = str(tmp_path / "accumulated.xlsx")
    _make_accumulated_report(acc_path)
    save_jira_settings("https://example.atlassian.net", "me@example.com", "token123")

    fm = FieldMapping(email="Email_Address", first_name="First_Name", last_name="Last_Name",
                       company="Company_Name", cid="CID")
    profile = ClientProfile(
        name="Test Client",
        accumulated_report_path=acc_path,
        field_mapping=fm,
        duplicate=DuplicateConfig(enabled=True),
        jira_ticket_key="PROJ-1234",
        client_mode="Lead QA & Upload",
    )
    save_profile(profile, get_clients_dir())

    new_leads = pd.DataFrame([
        {"Email_Address": "bob@new.com", "First_Name": "Bob", "Last_Name": "Lee", "Company_Name": "Beta", "CID": "1"},
    ])
    result = PipelineResult(valid_indices=[0], refund_reasons={})

    at = AppTest.from_file(_PAGE_PATH, default_timeout=15)
    at.session_state["run_new_leads"] = new_leads
    at.session_state["run_result"] = result
    at.session_state["run_result_for"] = "Test Client"
    at.run()

    finalize_button = next(b for b in at.button if b.label == "Finalize")
    finalize_button.click().run()
    assert not at.exception

    assert not any(c.key == "jira_link_Lead Report" for c in at.checkbox)
    assert any(c.key == "jira_link_Accumulated File" for c in at.checkbox)


def test_multi_tab_routes_different_cids_to_completely_different_files(tmp_path, monkeypatch):
    # End-to-end regression test for per-CID Lead Template files: some CID
    # groups go to a totally different workbook, not just another tab in
    # the same one.
    monkeypatch.chdir(tmp_path)
    acc_path = str(tmp_path / "accumulated.xlsx")
    _make_accumulated_report(acc_path)

    shared_path = str(tmp_path / "shared_template.xlsx")
    wb = openpyxl.Workbook()
    apac = wb.active
    apac.title = "APAC"
    apac.append(["Email_Address", "First_Name", "Last_Name", "Company_Name", "CID"])
    wb.save(shared_path)

    emea_only_path = str(tmp_path / "emea_only.xlsx")
    wb2 = openpyxl.Workbook()
    emea = wb2.active
    emea.title = "EMEA"
    emea.append(["Email_Address", "First_Name", "Last_Name", "Company_Name", "CID"])
    wb2.save(emea_only_path)

    fm = FieldMapping(email="Email_Address", first_name="First_Name", last_name="Last_Name",
                       company="Company_Name", cid="CID")
    profile = ClientProfile(
        name="Test Client",
        accumulated_report_path=acc_path,
        field_mapping=fm,
        client_mode="Lead QA",
        lead_template_path=shared_path,
        lead_template_multi_tab=True,
        lead_template_tabs=[
            LeadTemplateTab(sheet_name="APAC", cids=["1"]),  # blank file_path -> shared_path
            LeadTemplateTab(sheet_name="EMEA", cids=["2"], file_path=emea_only_path),
        ],
    )
    save_profile(profile, get_clients_dir())

    new_leads = pd.DataFrame([
        {"Email_Address": "apac@x.com", "First_Name": "A", "Last_Name": "One", "Company_Name": "X", "CID": "1"},
        {"Email_Address": "emea@x.com", "First_Name": "E", "Last_Name": "Two", "Company_Name": "X", "CID": "2"},
    ])
    result = PipelineResult(valid_indices=[0, 1], refund_reasons={})

    at = AppTest.from_file(_PAGE_PATH, default_timeout=15)
    at.session_state["run_new_leads"] = new_leads
    at.session_state["run_result"] = result
    at.session_state["run_result_for"] = "Test Client"
    at.run()

    finalize_button = next(b for b in at.button if b.label == "Finalize")
    finalize_button.click().run()
    assert not at.exception

    wb_shared = openpyxl.load_workbook(shared_path)
    assert wb_shared["APAC"].cell(row=2, column=1).value == "apac@x.com"

    wb_emea = openpyxl.load_workbook(emea_only_path)
    assert wb_emea["EMEA"].cell(row=2, column=1).value == "emea@x.com"

    # The EMEA-only file must not have gained an APAC lead, and vice versa.
    assert wb_shared["APAC"].max_row == 2
    assert wb_emea["EMEA"].max_row == 2


def test_jira_summary_uses_per_tab_sharepoint_links_for_multiple_lead_template_files(tmp_path, monkeypatch):
    # Regression test: per-CID Lead Template file routing means a single
    # run can write to more than one Lead Template workbook, each with its
    # own SharePoint link — the Jira link picker must offer one "Lead
    # Report" checkbox per distinct file, each pointing at that file's own
    # configured link rather than a single shared one.
    monkeypatch.chdir(tmp_path)
    acc_path = str(tmp_path / "accumulated.xlsx")
    _make_accumulated_report(acc_path)
    save_jira_settings("https://example.atlassian.net", "me@example.com", "token123")

    shared_path = str(tmp_path / "shared_template.xlsx")
    wb = openpyxl.Workbook()
    apac = wb.active
    apac.title = "APAC"
    apac.append(["Email_Address", "First_Name", "Last_Name", "Company_Name", "CID"])
    wb.save(shared_path)

    emea_only_path = str(tmp_path / "emea_only.xlsx")
    wb2 = openpyxl.Workbook()
    emea = wb2.active
    emea.title = "EMEA"
    emea.append(["Email_Address", "First_Name", "Last_Name", "Company_Name", "CID"])
    wb2.save(emea_only_path)

    fm = FieldMapping(email="Email_Address", first_name="First_Name", last_name="Last_Name",
                       company="Company_Name", cid="CID")
    profile = ClientProfile(
        name="Test Client",
        accumulated_report_path=acc_path,
        accumulated_report_link="https://madlog.sharepoint.com/:x:/s/Team/AccLink",
        field_mapping=fm,
        jira_ticket_key="PROJ-1234",
        client_mode="Lead QA",
        lead_template_path=shared_path,
        lead_template_link="https://madlog.sharepoint.com/:x:/s/Team/SharedLink",
        lead_template_multi_tab=True,
        lead_template_tabs=[
            LeadTemplateTab(sheet_name="APAC", cids=["1"]),  # blank link -> shared link
            LeadTemplateTab(sheet_name="EMEA", cids=["2"], file_path=emea_only_path,
                             link="https://madlog.sharepoint.com/:x:/s/Team/EmeaLink"),
        ],
    )
    save_profile(profile, get_clients_dir())

    new_leads = pd.DataFrame([
        {"Email_Address": "apac@x.com", "First_Name": "A", "Last_Name": "One", "Company_Name": "X", "CID": "1"},
        {"Email_Address": "emea@x.com", "First_Name": "E", "Last_Name": "Two", "Company_Name": "X", "CID": "2"},
    ])
    result = PipelineResult(valid_indices=[0, 1], refund_reasons={})

    at = AppTest.from_file(_PAGE_PATH, default_timeout=15)
    at.session_state["run_new_leads"] = new_leads
    at.session_state["run_result"] = result
    at.session_state["run_result_for"] = "Test Client"
    at.run()

    finalize_button = next(b for b in at.button if b.label == "Finalize")
    finalize_button.click().run()
    assert not at.exception

    checkbox_labels = [c.label for c in at.checkbox if c.key and c.key.startswith("jira_link_")]
    assert any("AccLink" in label for label in checkbox_labels)
    assert any("SharedLink" in label for label in checkbox_labels)
    assert any("EmeaLink" in label for label in checkbox_labels)


def test_complex_account_two_stage_finalize_previews_then_writes(tmp_path, monkeypatch):
    # End-to-end regression test for the Complex Account two-stage Finalize:
    # "Finalize (fill columns)" must not write anything, only preview the
    # column-filling rules on the valid leads — the Accumulated Report only
    # actually gets updated after "Confirm & Write".
    monkeypatch.chdir(tmp_path)
    acc_path = str(tmp_path / "accumulated.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Accumulated"
    ws.append(["Email", "First", "Last", "Company", "CID", "Capture Date", "Email Opt-in", "Business Phone"])
    wb.create_sheet("Refund").append(
        ["Email", "First", "Last", "Company", "CID", "Capture Date", "Email Opt-in", "Business Phone"])
    wb.save(acc_path)

    fm = FieldMapping(email="Email", first_name="First", last_name="Last", company="Company", cid="CID")
    profile = ClientProfile(
        name="Test Client",
        accumulated_report_path=acc_path,
        field_mapping=fm,
        complex_account=ComplexAccountConfig(enabled=True),
    )
    save_profile(profile, get_clients_dir())

    new_leads = pd.DataFrame([
        {"Email": "a@wipro.com", "First": "A", "Last": "One", "Company": "Wipro", "CID": "1",
         "Capture Date": "08/17/2026", "Email Opt-in": "Yes, Yes", "Business Phone": 919819719038},
    ])
    result = PipelineResult(valid_indices=[0], refund_reasons={})

    at = AppTest.from_file(_PAGE_PATH, default_timeout=15)
    at.session_state["run_new_leads"] = new_leads
    at.session_state["run_result"] = result
    at.session_state["run_result_for"] = "Test Client"
    at.run()
    assert not at.exception

    fill_button = next(b for b in at.button if b.label == "Finalize (fill columns)")
    fill_button.click().run()
    assert not at.exception

    # Nothing written yet — the Accumulated tab must still be just the header row.
    wb_after_fill = openpyxl.load_workbook(acc_path)
    assert wb_after_fill["Accumulated"].max_row == 1

    assert any("Preview: filled columns" in s.value for s in at.subheader)

    confirm_button = next(b for b in at.button if b.label == "Confirm & Write")
    confirm_button.click().run()
    assert not at.exception

    wb_final = openpyxl.load_workbook(acc_path)
    row = next(wb_final["Accumulated"].iter_rows(min_row=2, max_row=2, values_only=True))
    headers = next(wb_final["Accumulated"].iter_rows(min_row=1, max_row=1, values_only=True))
    written = dict(zip(headers, row))
    # A real date value now, not text — so Excel stores/filters it as a date.
    assert written["Capture Date"] == datetime.datetime(2026, 8, 17)
    assert written["Email Opt-in"] == "Yes"
    assert written["Business Phone"] == "91 9819719038"


def test_completed_checks_status_shown_for_enabled_checks_only(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    acc_path = str(tmp_path / "accumulated.xlsx")
    _make_accumulated_report(acc_path)

    fm = FieldMapping(email="Email_Address", first_name="First_Name", last_name="Last_Name",
                       company="Company_Name", cid="CID")
    profile = ClientProfile(
        name="Test Client", accumulated_report_path=acc_path, field_mapping=fm,
        duplicate=DuplicateConfig(enabled=True),
    )
    save_profile(profile, get_clients_dir())

    new_leads = pd.DataFrame([
        {"Email_Address": "a@x.com", "First_Name": "A", "Last_Name": "One", "Company_Name": "X", "CID": "1"},
    ])
    result = PipelineResult(valid_indices=[0], refund_reasons={})

    at = AppTest.from_file(_PAGE_PATH, default_timeout=15)
    at.session_state["run_new_leads"] = new_leads
    at.session_state["run_result"] = result
    at.session_state["run_result_for"] = "Test Client"
    at.run()
    assert not at.exception

    status_captions = [c.value for c in at.caption if "completed" in c.value]
    assert len(status_captions) == 1
    assert "Duplicate" in status_captions[0]
    assert "Leadcap" not in status_captions[0]  # not enabled for this client


def test_complex_account_flags_asset_url_mismatch_for_review_not_autocorrect(tmp_path, monkeypatch):
    # Regression test: Asset URN/Form URL/Dell Asset URL are already filled
    # in the leadfile — the tool must only flag a mismatch against the
    # specifications file for review, never silently rewrite them.
    monkeypatch.chdir(tmp_path)
    acc_path = str(tmp_path / "accumulated.xlsx")
    _make_accumulated_report(acc_path)

    specs_path = str(tmp_path / "specs.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Asset Name", "URN ", "Asset URL 1", "Asset URL 2", "Dell URL"])
    ws.append(["Fuel AI Innovation", "DT2503G0007_033", "https://a.com/1", "https://a.com/2", "https://dell.com/x"])
    wb.save(specs_path)

    fm = FieldMapping(email="Email_Address", first_name="First_Name", last_name="Last_Name",
                       company="Company_Name", cid="CID")
    profile = ClientProfile(
        name="Test Client", accumulated_report_path=acc_path, field_mapping=fm,
        complex_account=ComplexAccountConfig(enabled=True, specifications_path=specs_path),
    )
    save_profile(profile, get_clients_dir())

    new_leads = pd.DataFrame([
        {"Email_Address": "a@x.com", "First_Name": "A", "Last_Name": "One", "Company_Name": "X", "CID": "1",
         "Asset Title": "Fuel AI Innovation", "Asset URN": "WRONG_URN",
         "Form URL": "https://a.com/1", "Dell Asset URL": "https://dell.com/x"},
    ])

    # This exercises the same functions pages/2_Run_Check.py's "Run Check"
    # button calls: load the specs file, run the Complex Account checks
    # alongside the normal pipeline, and merge the results together.
    from core.complex_account import (
        check_complex_account_conditions, merge_complex_account_review, load_asset_specifications,
    )

    accumulated_leads = pd.read_excel(acc_path, sheet_name="Accumulated")
    asset_specs = load_asset_specifications(specs_path)
    complex_review = check_complex_account_conditions(new_leads, asset_specs)
    result = run_pipeline(new_leads, profile, accumulated_leads, {}, [])
    merge_complex_account_review(result, complex_review)

    assert 0 not in result.valid_indices
    assert 0 in result.review_reasons
    assert any("Asset URN" in str(d) for d in result.review_reasons[0])
