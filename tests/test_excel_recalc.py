import os
from unittest.mock import patch

import openpyxl
import pytest

from core.excel_recalc import recalculate_workbook

try:
    import win32com.client  # noqa: F401
    _PYWIN32_AVAILABLE = True
except ImportError:
    _PYWIN32_AVAILABLE = False


@pytest.mark.skipif(not _PYWIN32_AVAILABLE, reason="pywin32/Excel not available on this machine")
def test_recalculate_workbook_produces_fresh_cached_formula_value(tmp_path):
    # openpyxl never evaluates formulas — a formula cell it just wrote has no
    # cached value at all until something (normally Excel) actually opens and
    # calculates the file. This is the exact mechanism behind the reported
    # bug: a Pacing Overview column driven by a formula read back as blank
    # right after this app appended new leads via openpyxl, even though the
    # underlying data was already correct.
    path = str(tmp_path / "formulas.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = 5
    ws["A2"] = 10
    ws["B1"] = "=A1+A2"
    wb.save(path)

    before = openpyxl.load_workbook(path, data_only=True)
    assert before["Sheet"]["B1"].value is None

    recalculated_path = recalculate_workbook(path)
    assert recalculated_path != path  # a temp copy, not the original file

    after = openpyxl.load_workbook(recalculated_path, data_only=True)
    assert after["Sheet"]["B1"].value == 15

    # The original file must never be touched — only a temp copy is opened.
    untouched = openpyxl.load_workbook(path, data_only=True)
    assert untouched["Sheet"]["B1"].value is None


def test_recalculate_workbook_falls_back_to_original_path_when_com_fails(tmp_path):
    path = str(tmp_path / "formulas.xlsx")
    wb = openpyxl.Workbook()
    wb.active["A1"] = 1
    wb.save(path)

    with patch("win32com.client.DispatchEx", side_effect=RuntimeError("Excel automation is unavailable")):
        result = recalculate_workbook(path, timeout_seconds=5)

    assert result == path


def test_recalculate_workbook_returns_original_path_when_pywin32_missing(tmp_path):
    path = str(tmp_path / "formulas.xlsx")
    wb = openpyxl.Workbook()
    wb.active["A1"] = 1
    wb.save(path)

    real_import = __import__

    def _fake_import(name, *args, **kwargs):
        if name in ("pythoncom", "win32com.client", "win32process"):
            raise ImportError(f"No module named '{name}'")
        return real_import(name, *args, **kwargs)

    with patch("builtins.__import__", side_effect=_fake_import):
        result = recalculate_workbook(path)

    assert result == path
