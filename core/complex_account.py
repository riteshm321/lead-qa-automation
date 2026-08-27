import datetime
import io
import re

import openpyxl
import pandas as pd

from core.check_result import ReviewDetail
from core.matching import extract_domain

# Hardcoded to Dell APAC's actual column names (per design decision: not
# worth a full mapping UI for a single client). If a future complex account
# needs different names, add configurability then.
COUNTRY_COLUMN = "Country"
ACCOUNT_ID_COLUMN = "Account ID"
COMPANY_COLUMN = "Company"
CAPTURE_DATE_COLUMN = "Capture Date"
EMAIL_OPTIN_COLUMN = "Email Opt-in"
PHONE_COLUMN = "Business Phone"
ASSET_TITLE_COLUMN = "Asset Title"
ASSET_URN_COLUMN = "Asset URN"
FORM_URL_COLUMN = "Form URL"
DELL_ASSET_URL_COLUMN = "Dell Asset URL"
TOP_TOPICS_COLUMN = "Additional Data Point (poll questions, dynamic data, etc)  1"
INSTALLED_TECH_COLUMN = "Additional Data Point (poll questions, dynamic data, etc)  2"
PBS_COLUMN = "Additional Data Point (poll questions, dynamic data, etc)  3"
DOWNLOAD_DAY_COLUMN = "Asset download day"
DOWNLOAD_MONTH_COLUMN = "Asset download month"
DOWNLOAD_YEAR_COLUMN = "Asset download year"
DOWNLOAD_YEAR_VALUE = 2026
AGREED_CONTACTED_COLUMN = "Agreed to be contacted by Dell Technologies"
PHONE_OPTIN_COLUMN = "Phone Opt-In"
MAIL_OPTIN_COLUMN = "Mail Opt-In"
SIGNAL_NOTES_COLUMN = "Signal Notes"
# Per-CID constant, not derived from anything in the leadfile -- confirmed
# business rule for this client's two CID groups.
AGREED_CONTACTED_BY_CID = {"119414": "No", "119415": "Yes"}
# Which specifications-file link column Form URL is checked/corrected
# against, per CID -- "Publisher Link [AU]_BHRS" is for 119415 (AU),
# "Publisher Link INDIA]_ECS" is for 119414 (India); each CID has its own
# link, not two interchangeable options. A CID with no entry here gets no
# Form URL check at all -- no rule was specified for any other CID.
FORM_URL_SPEC_KEY_BY_CID = {"119414": "india_link", "119415": "au_link"}

_KNOWN_COLUMNS = (
    COUNTRY_COLUMN, ACCOUNT_ID_COLUMN, COMPANY_COLUMN, CAPTURE_DATE_COLUMN,
    EMAIL_OPTIN_COLUMN, PHONE_COLUMN, ASSET_TITLE_COLUMN, ASSET_URN_COLUMN,
    FORM_URL_COLUMN, DELL_ASSET_URL_COLUMN, TOP_TOPICS_COLUMN,
    INSTALLED_TECH_COLUMN, PBS_COLUMN, DOWNLOAD_DAY_COLUMN,
    DOWNLOAD_MONTH_COLUMN, DOWNLOAD_YEAR_COLUMN, AGREED_CONTACTED_COLUMN,
    PHONE_OPTIN_COLUMN, MAIL_OPTIN_COLUMN, SIGNAL_NOTES_COLUMN,
)


def _normalize_header_text(value) -> str:
    # Same normalization as core/excel_io.py's Lead Template column
    # matching: strips all non-alphanumeric characters so "Capture Date",
    # "CaptureDate" and "capture_date" all match -- real leadfile exports
    # frequently drop or rename separators in headers.
    return re.sub(r"[^a-z0-9]", "", str(value).strip().lower())


def _normalize_known_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Renames any column matching one of this module's hardcoded names
    (ignoring case/spacing) to that exact name, so the exact-match column
    lookups below still fire when a leadfile's real header text (e.g.
    "CaptureDate") differs only cosmetically from what's hardcoded.
    """
    lookup = {_normalize_header_text(name): name for name in _KNOWN_COLUMNS}
    rename = {
        col: lookup[_normalize_header_text(col)]
        for col in df.columns
        if _normalize_header_text(col) in lookup and col != lookup[_normalize_header_text(col)]
    }
    return df.rename(columns=rename) if rename else df


def _norm_domain(value) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    return "" if text in ("", "nan") else text


def _norm_cid(value) -> str:
    # A CID column with even one blank cell elsewhere gets silently upcast
    # by pandas from int64 to float64, turning every value from e.g. 119414
    # into 119414.0 -- without normalizing this, a per-CID rule keyed on
    # the clean digit string ("119414") would never match.
    text = str(value).strip() if value is not None else ""
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def load_tal_index(tal_path: str) -> dict[str, list[dict]]:
    """Loads the (large, ~500k+ row) TAL reference file into a
    domain -> [{"account_id", "account_name", "country_code"}, ...] index,
    reading only the 4 columns actually needed to keep memory/load time
    reasonable. More than one TAL row can share the same domain (different,
    genuinely distinct accounts, not just duplicate rows) — match_tal_account
    resolves that ambiguity per lead using the lead's own country.
    """
    df = pd.read_csv(tal_path, usecols=["web_domain", "account_id", "account_name", "country_code"])
    index: dict[str, list[dict]] = {}
    for row in df.itertuples(index=False):
        domain = _norm_domain(row.web_domain)
        if not domain:
            continue
        index.setdefault(domain, []).append({
            "account_id": row.account_id,
            "account_name": row.account_name,
            "country_code": str(row.country_code or "").strip().upper(),
        })
    return index


def match_tal_account(domain: str, country: str, tal_index: dict[str, list[dict]]) -> tuple[str | None, str | None]:
    """Returns (account_id, account_name) for the given domain, or (None, None)
    if the domain isn't in the TAL at all. When a domain maps to more than one
    distinct account, prefers one whose country_code matches the lead's own
    Country — if that still doesn't resolve it, returns the first candidate
    rather than leaving it blank (a real client-facing report should never
    show an empty Account ID just because two TAL rows share a domain).
    """
    candidates = tal_index.get(_norm_domain(domain))
    if not candidates:
        return None, None
    if len(candidates) > 1:
        country_norm = str(country or "").strip().upper()
        if country_norm:
            for candidate in candidates:
                if candidate["country_code"] == country_norm:
                    return candidate["account_id"], candidate["account_name"]
    chosen = candidates[0]
    return chosen["account_id"], chosen["account_name"]


def apply_tal_mapping(
    leads_df: pd.DataFrame, email_column: str, country_column: str,
    account_id_column: str, company_column: str, tal_index: dict[str, list[dict]],
) -> pd.DataFrame:
    """Fills account_id_column from the TAL for every lead whose email
    domain matches, replacing company_column with the TAL's own company
    name for those leads. A lead with no TAL match gets a blank account id
    and keeps its original company name untouched.
    """
    df = leads_df.copy()
    account_ids = []
    companies = list(df[company_column]) if company_column in df.columns else [""] * len(df)
    for i, (_, row) in enumerate(df.iterrows()):
        domain = extract_domain(row.get(email_column))
        account_id, account_name = match_tal_account(domain, row.get(country_column), tal_index)
        account_ids.append(account_id or "")
        if account_name:
            companies[i] = account_name
    df[account_id_column] = account_ids
    df[company_column] = companies
    return df


def _find_csv_header_row(text: str, required_column: str, max_scan: int = 15) -> int:
    # These exports carry a couple of "Client:"/"Program:" metadata lines
    # (and a blank line) above the real header row.
    required_norm = required_column.strip().lower()
    for i, line in enumerate(text.splitlines()[:max_scan]):
        if required_norm in line.strip().lower():
            return i
    return 0


def load_domain_value_map(
    file_obj, domain_column: str, value_column: str,
    aggregate: bool = False, skip_values: set[str] | None = None,
) -> dict[str, str]:
    """Reads a CID-specific reference export (Installed Technologies or
    Predictive Buying Stage) into a domain -> value dict. file_obj is
    anything with .read() returning bytes (a Streamlit UploadedFile or a
    plain open file).

    aggregate: a domain can appear on more than one row (e.g. Installed
    Technologies lists one technology per row) — when True, every distinct
    value seen for a domain is joined with ", " in first-seen order,
    instead of only the last row winning.

    skip_values: values to treat as "nothing to report" (case-insensitive
    exact match) — e.g. Predictive Buying Stage's "No Active Signals"
    should leave that domain unmapped rather than showing the label text
    itself.
    """
    raw = file_obj.read()
    if isinstance(raw, bytes):
        text = raw.decode("utf-8-sig", errors="replace")
    else:
        text = raw
    header_row = _find_csv_header_row(text, domain_column)
    df = pd.read_csv(io.StringIO(text), skiprows=header_row)
    skip_norm = {s.strip().lower() for s in (skip_values or ())}

    mapping: dict[str, str] = {}
    seen_values: dict[str, list[str]] = {}
    for _, row in df.iterrows():
        domain = _norm_domain(row.get(domain_column))
        if not domain:
            continue
        value = row.get(value_column)
        if value is None or not str(value).strip() or str(value).strip().lower() == "nan":
            continue
        value_text = str(value).strip()
        if value_text.lower() in skip_norm:
            continue
        if aggregate:
            values = seen_values.setdefault(domain, [])
            if value_text not in values:
                values.append(value_text)
            mapping[domain] = ", ".join(values)
        else:
            mapping[domain] = value_text
    return mapping


_DATE_FORMATS = (
    "%m/%d/%Y", "%m/%d/%y", "%d-%b-%Y", "%d-%B-%Y",
    "%B %d, %Y", "%b %d, %Y", "%d %B %Y", "%Y-%m-%d",
)


def reformat_capture_date(value) -> str | None:
    """Returns the date as mm/dd/yyyy text, or None if value is blank or
    couldn't be parsed as a date at all (caller flags that lead for review —
    per instruction, this should never actually be blank in practice).
    US-style m/d/y is tried first since that's this client's own convention
    (and the ambiguous case, e.g. "03/04/2026", only has one sane reading
    without more context).
    """
    if isinstance(value, (datetime.datetime, datetime.date)):
        return value.strftime("%m/%d/%Y")
    text = str(value).strip() if value is not None else ""
    if not text or text.lower() == "nan":
        return None
    for fmt in _DATE_FORMATS:
        try:
            return datetime.datetime.strptime(text, fmt).strftime("%m/%d/%Y")
        except ValueError:
            continue
    try:
        return pd.to_datetime(text).strftime("%m/%d/%Y")
    except (ValueError, TypeError):
        return None


def clean_email_optin(value) -> str | None:
    """Collapses a verbose opt-in value ("Yes, I would like Dell to contact
    me by email...") down to a bare "Yes"/"No". Returns None if the value
    doesn't clearly contain exactly one of "yes"/"no" (caller flags that
    lead for review rather than guessing).
    """
    text = str(value).strip().lower() if value is not None else ""
    has_yes = "yes" in text
    has_no = "no" in text
    if has_yes and not has_no:
        return "Yes"
    if has_no and not has_yes:
        return "No"
    return None


def asset_download_parts(capture_date) -> tuple[int, str]:
    """(day-of-month as a number, full month name) from a Capture Date
    value — either an already mm/dd/yyyy-formatted string, or a real
    date/datetime object. Returned as a number (not zero-padded text) so
    Excel stores and filters it as a number rather than text."""
    parsed = (
        datetime.datetime.strptime(capture_date, "%m/%d/%Y")
        if isinstance(capture_date, str) else capture_date
    )
    return parsed.day, parsed.strftime("%B")


def format_phone(value) -> str:
    """Strips every non-digit character, then inserts a single space after
    the first 2 digits (e.g. "+91-92-929-29292" -> "91 9292929292")."""
    digits = re.sub(r"\D", "", str(value)) if value is not None else ""
    return digits if len(digits) <= 2 else f"{digits[:2]} {digits[2:]}"


def load_asset_specifications(path: str) -> dict[str, dict]:
    """Reads the "Specifications Campaigns - BANT NTQ & EHS" workbook into
    normalized-Asset-Name -> {"urn", "au_link", "india_link", "dell_url"}.

    Header text on this sheet wraps onto a second line and carries
    bracketed "[to be filled in by ...]" annotations that shift over
    time (confirmed: the real file's headers changed between two
    sessions of this exact feature) -- matched by substring, not exact
    text, so a minor header edit doesn't silently break the whole thing.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        header_cells = next(ws.iter_rows(min_row=1, max_row=1))
        name_i = urn_i = au_link_i = india_link_i = dell_i = None
        for i, cell in enumerate(header_cells):
            if cell.value is None:
                continue
            header = str(cell.value).strip().lower()
            if "asset name" in header:
                name_i = i
            elif "urn" in header:
                urn_i = i
            elif "[au]" in header:
                au_link_i = i
            elif "india" in header:
                india_link_i = i
            elif "dell" in header:
                dell_i = i
        if name_i is None:
            raise ValueError(f"'{path}' has no 'Asset Name' column")

        specs: dict[str, dict] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = row[name_i] if name_i < len(row) else None
            if name is None or not str(name).strip():
                continue
            specs[str(name).strip().lower()] = {
                "urn": row[urn_i] if urn_i is not None and urn_i < len(row) else "",
                "au_link": row[au_link_i] if au_link_i is not None and au_link_i < len(row) else "",
                "india_link": row[india_link_i] if india_link_i is not None and india_link_i < len(row) else "",
                "dell_url": row[dell_i] if dell_i is not None and dell_i < len(row) else "",
            }
        return specs
    finally:
        wb.close()


def check_asset_url_mismatches(
    leads_df: pd.DataFrame, asset_specs: dict[str, dict], field_mapping=None,
) -> dict[int, list[ReviewDetail]]:
    """Flags leads whose already-filled Asset URN / Form URL / Dell Asset
    URL don't match what the specifications file says for that Asset
    Title. Flagged here for review; apply_complex_account_rules is what
    actually corrects the value (from the specifications file) once the
    lead is approved as valid -- this function only ever flags, never
    writes. An Asset Title not found in the specifications file at all is
    itself flagged, since none of the other three fields can be verified
    without it.

    Form URL is checked against a CID-specific column, not either of two
    interchangeable URLs -- see FORM_URL_SPEC_KEY_BY_CID. field_mapping is
    required to know a lead's CID for that check; pass None to skip the
    Form URL check for every lead (e.g. a caller that doesn't have a
    FieldMapping handy).
    """
    review: dict[int, list[ReviewDetail]] = {}
    if ASSET_TITLE_COLUMN not in leads_df.columns:
        return review
    for idx, row in leads_df.iterrows():
        asset_title = str(row.get(ASSET_TITLE_COLUMN, "") or "").strip()
        spec = asset_specs.get(asset_title.lower())
        if spec is None:
            if asset_title:
                review.setdefault(idx, []).append(ReviewDetail(
                    check="Complex Account",
                    message="Asset Title not found in the specifications file -- Asset URN/Dell Asset "
                            "URL/Form URL can't be verified",
                    lead_value=asset_title,
                ))
            continue

        context = f"specifications file entry for \"{asset_title}\""

        urn = str(row.get(ASSET_URN_COLUMN, "") or "").strip()
        expected_urn = str(spec["urn"]).strip()
        if urn != expected_urn:
            review.setdefault(idx, []).append(ReviewDetail(
                check="Complex Account",
                message="Asset URN doesn't match the specifications file (will be corrected automatically)",
                lead_value=urn, candidate_value=expected_urn, candidate_context=context,
            ))

        dell_url = str(row.get(DELL_ASSET_URL_COLUMN, "") or "").strip()
        expected_dell_url = str(spec["dell_url"]).strip()
        if dell_url != expected_dell_url:
            review.setdefault(idx, []).append(ReviewDetail(
                check="Complex Account",
                message="Dell Asset URL doesn't match the specifications file (will be corrected automatically)",
                lead_value=dell_url, candidate_value=expected_dell_url, candidate_context=context,
            ))

        if field_mapping is not None:
            spec_key = FORM_URL_SPEC_KEY_BY_CID.get(_norm_cid(row.get(field_mapping.cid, "")))
            if spec_key is not None:
                form_url = str(row.get(FORM_URL_COLUMN, "") or "").strip()
                expected_form_url = str(spec.get(spec_key, "")).strip()
                if form_url != expected_form_url:
                    review.setdefault(idx, []).append(ReviewDetail(
                        check="Complex Account",
                        message="Form URL doesn't match the specifications file for this CID "
                                "(will be corrected automatically)",
                        lead_value=form_url, candidate_value=expected_form_url, candidate_context=context,
                    ))
    return review


def check_complex_account_conditions(
    leads_df: pd.DataFrame, asset_specs: dict[str, dict] | None = None, field_mapping=None,
) -> dict[int, list[ReviewDetail]]:
    """Evaluates the Complex Account conditions that can actually flag a
    lead — a Capture Date that's blank/unparseable, an Email Opt-in value
    that isn't clearly Yes/No, or an already-filled Asset URN/Form
    URL/Dell Asset URL that doesn't match the specifications file for that
    Asset Title — without touching any column.

    Used at Run Check time, before the valid/refund/review split, so these
    leads get resolved through the same Refund/Needs Review flow as every
    other check. The column-filling rules (TAL mapping, Installed
    Technologies/Predictive Buying Stage, phone/date formatting — see
    apply_complex_account_rules) are deliberately deferred to a separate
    step run only on the leads that end up valid, since there's no point
    enriching a lead that's about to be refunded.
    """
    leads_df = _normalize_known_columns(leads_df)
    review: dict[int, list[ReviewDetail]] = {}
    if CAPTURE_DATE_COLUMN in leads_df.columns:
        for idx, row in leads_df.iterrows():
            if reformat_capture_date(row.get(CAPTURE_DATE_COLUMN)) is None:
                review.setdefault(idx, []).append(ReviewDetail(
                    check="Complex Account", message="Capture Date is blank or unparseable",
                    lead_value=str(row.get(CAPTURE_DATE_COLUMN, "")),
                ))
    if EMAIL_OPTIN_COLUMN in leads_df.columns:
        for idx, row in leads_df.iterrows():
            if clean_email_optin(row.get(EMAIL_OPTIN_COLUMN)) is None:
                review.setdefault(idx, []).append(ReviewDetail(
                    check="Complex Account", message="Email Opt-in value is not clearly Yes/No",
                    lead_value=str(row.get(EMAIL_OPTIN_COLUMN, "")),
                ))
    if asset_specs is not None:
        for idx, details in check_asset_url_mismatches(leads_df, asset_specs, field_mapping).items():
            review.setdefault(idx, []).extend(details)
    return review


def apply_complex_account_rules(
    leads_df: pd.DataFrame,
    field_mapping,
    tal_index: dict[str, list[dict]] | None,
    installed_tech_map: dict[str, str],
    pbs_map: dict[str, str],
    asset_specs: dict[str, dict] | None = None,
) -> tuple[pd.DataFrame, dict[int, list[ReviewDetail]], dict[int, list[str]]]:
    """Applies every Complex Account column-filling rule to a copy of
    leads_df and returns (enriched_df, review_reasons, corrections).
    review_reasons only ever contains entries for the two rules that can't
    safely auto-decide (Capture Date, Email Opt-in); every other rule
    always produces a value (TAL: blank Account ID on no match; Installed
    Technologies/Predictive Buying Stage: blank on no file/no match).
    Values are ReviewDetail objects, matching the shape every other check
    in core/checks/ returns, so they render in the same "Needs Review" UI
    unchanged.

    corrections: {row index: ["Asset URN: \"old\" -> \"new\"", ...]} for
    every Asset URN/Dell Asset URL/Form URL value this function silently
    overwrote with the specifications file's value. A lead flagged by
    check_asset_url_mismatches and then approved as valid gets corrected
    here rather than written through with its original wrong value --
    callers should surface this to the user (e.g. "N lead(s) had values
    corrected") since it happens without further confirmation.

    installed_tech_map / pbs_map: {domain: value}, covering every CID in
    one file (not split per CID) -- a domain missing from the map (no file
    uploaded, or no match for that domain) gets that lead's corresponding
    column cleared to blank, per design.

    Agreed to be contacted by Dell Technologies / Phone Opt-In: fixed
    business rules for this client's two CID groups (see
    AGREED_CONTACTED_BY_CID), not derived from the leadfile at all. Mail
    Opt-In and Signal Notes are always cleared to blank, regardless of
    what the leadfile has.
    """
    df = _normalize_known_columns(leads_df.copy())
    review: dict[int, list[ReviewDetail]] = {}
    corrections: dict[int, list[str]] = {}

    if tal_index is not None:
        df = apply_tal_mapping(df, field_mapping.email, COUNTRY_COLUMN, ACCOUNT_ID_COLUMN, COMPANY_COLUMN, tal_index)

    for idx, row in df.iterrows():
        domain = _norm_domain(extract_domain(row.get(field_mapping.email)))

        it_value = installed_tech_map.get(domain)
        if INSTALLED_TECH_COLUMN in df.columns:
            df.at[idx, INSTALLED_TECH_COLUMN] = f"Installed Technologies: {it_value}" if it_value else ""

        pbs_value = pbs_map.get(domain)
        if PBS_COLUMN in df.columns:
            df.at[idx, PBS_COLUMN] = f"Predictive Buying Stage: {pbs_value}" if pbs_value else ""

        cid = _norm_cid(row.get(field_mapping.cid, ""))
        if AGREED_CONTACTED_COLUMN in df.columns:
            df.at[idx, AGREED_CONTACTED_COLUMN] = AGREED_CONTACTED_BY_CID.get(cid, "")

        if asset_specs is not None and ASSET_TITLE_COLUMN in df.columns:
            spec = asset_specs.get(str(row.get(ASSET_TITLE_COLUMN, "") or "").strip().lower())
            if spec is not None:
                if ASSET_URN_COLUMN in df.columns:
                    current = str(row.get(ASSET_URN_COLUMN, "") or "").strip()
                    expected = str(spec["urn"]).strip()
                    if current != expected:
                        df.at[idx, ASSET_URN_COLUMN] = expected
                        corrections.setdefault(idx, []).append(f"Asset URN: \"{current}\" -> \"{expected}\"")
                if DELL_ASSET_URL_COLUMN in df.columns:
                    current = str(row.get(DELL_ASSET_URL_COLUMN, "") or "").strip()
                    expected = str(spec["dell_url"]).strip()
                    if current != expected:
                        df.at[idx, DELL_ASSET_URL_COLUMN] = expected
                        corrections.setdefault(idx, []).append(f"Dell Asset URL: \"{current}\" -> \"{expected}\"")
                spec_key = FORM_URL_SPEC_KEY_BY_CID.get(cid)
                if spec_key is not None and FORM_URL_COLUMN in df.columns:
                    current = str(row.get(FORM_URL_COLUMN, "") or "").strip()
                    expected = str(spec.get(spec_key, "")).strip()
                    if current != expected:
                        df.at[idx, FORM_URL_COLUMN] = expected
                        corrections.setdefault(idx, []).append(f"Form URL: \"{current}\" -> \"{expected}\"")

    if PHONE_OPTIN_COLUMN in df.columns:
        df[PHONE_OPTIN_COLUMN] = "Yes"

    if MAIL_OPTIN_COLUMN in df.columns:
        df[MAIL_OPTIN_COLUMN] = ""

    if SIGNAL_NOTES_COLUMN in df.columns:
        df[SIGNAL_NOTES_COLUMN] = ""

    if TOP_TOPICS_COLUMN in df.columns:
        df[TOP_TOPICS_COLUMN] = df[TOP_TOPICS_COLUMN].apply(
            lambda v: f"Top Trending Topics: {v}" if pd.notna(v) and str(v).strip() else v
        )

    capture_date_ok = pd.Series(True, index=df.index)
    if CAPTURE_DATE_COLUMN in df.columns:
        # A string-typed column (e.g. pandas' pyarrow-backed "str" dtype)
        # rejects assigning a real date object cell-by-cell below — widen it
        # to plain object dtype first so it can hold dates.
        df[CAPTURE_DATE_COLUMN] = df[CAPTURE_DATE_COLUMN].astype(object)
        for idx, row in df.iterrows():
            formatted = reformat_capture_date(row.get(CAPTURE_DATE_COLUMN))
            if formatted is None:
                review.setdefault(idx, []).append(ReviewDetail(
                    check="Complex Account", message="Capture Date is blank or unparseable",
                    lead_value=str(row.get(CAPTURE_DATE_COLUMN, "")),
                ))
                capture_date_ok[idx] = False
            else:
                # Store a real date, not text — so Excel writes/filters it as
                # an actual date instead of flagging "Number Stored as Text".
                df.at[idx, CAPTURE_DATE_COLUMN] = datetime.datetime.strptime(formatted, "%m/%d/%Y").date()

    if EMAIL_OPTIN_COLUMN in df.columns:
        for idx, row in df.iterrows():
            cleaned = clean_email_optin(row.get(EMAIL_OPTIN_COLUMN))
            if cleaned is None:
                review.setdefault(idx, []).append(ReviewDetail(
                    check="Complex Account", message="Email Opt-in value is not clearly Yes/No",
                    lead_value=str(row.get(EMAIL_OPTIN_COLUMN, "")),
                ))
            else:
                df.at[idx, EMAIL_OPTIN_COLUMN] = cleaned

    if CAPTURE_DATE_COLUMN in df.columns and DOWNLOAD_DAY_COLUMN in df.columns:
        # Same string-dtype widening as Capture Date above — day/year are
        # real numbers now, not zero-padded text, so Excel doesn't flag them.
        df[DOWNLOAD_DAY_COLUMN] = df[DOWNLOAD_DAY_COLUMN].astype(object)
        df[DOWNLOAD_YEAR_COLUMN] = df[DOWNLOAD_YEAR_COLUMN].astype(object)
        for idx, row in df.iterrows():
            if not capture_date_ok[idx]:
                continue
            day, month = asset_download_parts(row.get(CAPTURE_DATE_COLUMN))
            df.at[idx, DOWNLOAD_DAY_COLUMN] = day
            df.at[idx, DOWNLOAD_MONTH_COLUMN] = month
            df.at[idx, DOWNLOAD_YEAR_COLUMN] = DOWNLOAD_YEAR_VALUE

    if PHONE_COLUMN in df.columns:
        df[PHONE_COLUMN] = df[PHONE_COLUMN].apply(format_phone)

    return df, review, corrections


def merge_complex_account_review(result, complex_review: dict[int, list[ReviewDetail]]) -> None:
    """Merges Complex Account review flags into an existing PipelineResult
    in place, respecting the same fail > review > valid priority
    run_pipeline() itself uses — a lead already auto-refunded by one of the
    standard checks stays refunded; anything else moves to (or stays in)
    review.
    """
    for idx, reasons in complex_review.items():
        if idx in result.refund_reasons:
            continue
        result.review_reasons.setdefault(idx, []).extend(reasons)
        if idx in result.valid_indices:
            result.valid_indices.remove(idx)
