import csv
import datetime
import io
import os
import re
import shutil
import xml.etree.ElementTree as ET
import zipfile
from copy import copy
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.formula.translate import Translator
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries

from core.models import FieldMapping, LeadTemplateTab


def _read_external_link_parts(path: str) -> dict[str, bytes]:
    with zipfile.ZipFile(path, "r") as zin:
        return {n: zin.read(n) for n in zin.namelist() if n.startswith("xl/externalLinks/")}


def _restore_external_link_parts(path: str, original_parts: dict[str, bytes]) -> None:
    # openpyxl loses/corrupts the cached values in xl/externalLinks/*.xml when
    # it round-trips a workbook that references another (possibly closed)
    # workbook — e.g. a data-validation picklist backed by an external file.
    # Excel still opens the result, but flags it as needing repair every
    # time. Since we never touch external links ourselves, restoring the
    # original bytes for exactly those parts is always safe and correct.
    with zipfile.ZipFile(path, "r") as zin:
        entries = {n: zin.read(n) for n in zin.namelist()}
    entries.update(original_parts)
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, data in entries.items():
            zout.writestr(name, data)


_MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_PKG_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"


def _worksheet_xml_path_for_sheet(path: str, sheet_name: str) -> str | None:
    """The zip entry name (e.g. "xl/worksheets/sheet4.xml") holding one
    sheet's XML, resolved the same way Excel itself does: sheet name ->
    r:id in xl/workbook.xml -> target file in xl/_rels/workbook.xml.rels.
    Returns None if the sheet or either mapping file can't be found."""
    with zipfile.ZipFile(path, "r") as zin:
        names = zin.namelist()
        if "xl/workbook.xml" not in names or "xl/_rels/workbook.xml.rels" not in names:
            return None
        workbook_xml = zin.read("xl/workbook.xml")
        rels_xml = zin.read("xl/_rels/workbook.xml.rels")

    sheets_el = ET.fromstring(workbook_xml).find(f"{{{_MAIN_NS}}}sheets")
    if sheets_el is None:
        return None
    rid = None
    for sheet_el in sheets_el:
        if sheet_el.get("name") == sheet_name:
            rid = sheet_el.get(f"{{{_REL_NS}}}id")
            break
    if rid is None:
        return None

    for rel_el in ET.fromstring(rels_xml):
        if rel_el.get("Id") == rid:
            target = rel_el.get("Target", "")
            # Relative to the xl/ folder ("worksheets/sheet4.xml") in a
            # freshly-authored file, but openpyxl itself writes an
            # absolute in-package path ("/xl/worksheets/sheet4.xml") when
            # it re-saves -- handle whichever form is present.
            return target.lstrip("/") if target.startswith("/") else f"xl/{target}"
    return None


def _read_worksheet_ext_list(path: str, sheet_name: str) -> bytes | None:
    """Raw <extLst>...</extLst> bytes from one worksheet's XML, if present.

    This is where Excel 2010+ extended features live -- x14 conditional
    formatting (e.g. highlighting a cell whose value isn't in an allowed
    picklist) and x14 data validation (e.g. a dropdown whose options
    depend on another cell's value). openpyxl doesn't model either and
    silently drops them on every save -- confirmed by re-opening a
    real Lead Template that had already been through this tool once:
    loading the untouched original template warns "Data Validation
    extension is not supported and will be removed"; loading the
    already-saved file raises no such warning, because it's already gone.
    Capturing the bytes here lets them be spliced back in afterward.
    """
    sheet_xml_path = _worksheet_xml_path_for_sheet(path, sheet_name)
    if sheet_xml_path is None:
        return None
    with zipfile.ZipFile(path, "r") as zin:
        if sheet_xml_path not in zin.namelist():
            return None
        xml = zin.read(sheet_xml_path).decode("utf-8")
    start = xml.find("<extLst>")
    if start == -1:
        return None
    end = xml.find("</extLst>", start)
    if end == -1:
        return None
    ext_list = xml[start:end + len("</extLst>")]

    # Content inside <extLst> can reference namespace prefixes (e.g. an
    # xr:uid="..." attribute Excel stamps on x14:dataValidation) that are
    # declared on the ORIGINAL file's root <worksheet> tag, not repeated
    # locally within <extLst> itself. Splicing the fragment as-is into a
    # different file's worksheet XML (whose own root tag won't necessarily
    # declare the same namespaces) then fails to parse with an "unbound
    # prefix" error. Promote every such declaration onto <extLst> itself
    # so the fragment is self-contained regardless of where it lands.
    root_tag_start = xml.find("<worksheet")
    root_tag = xml[root_tag_start:xml.find(">", root_tag_start) + 1]
    root_namespaces = dict(re.findall(r'xmlns:(\w+)="([^"]+)"', root_tag))
    used_prefixes = set(re.findall(r'[<\s](\w+):\w', ext_list))
    declared_prefixes = set(re.findall(r'xmlns:(\w+)=', ext_list))
    missing = sorted(p for p in used_prefixes - declared_prefixes if p in root_namespaces)
    if missing:
        extra_ns = "".join(f' xmlns:{p}="{root_namespaces[p]}"' for p in missing)
        ext_list = ext_list.replace("<extLst>", f"<extLst{extra_ns}>", 1)

    return ext_list.encode("utf-8")


def _restore_worksheet_ext_list(path: str, sheet_name: str, ext_list_xml: bytes) -> None:
    # <extLst> is always the last child of the <worksheet> root element per
    # the OOXML schema, so splicing it back in right before the closing tag
    # reproduces exactly where openpyxl would have written its own (had it
    # understood the content well enough to keep it).
    sheet_xml_path = _worksheet_xml_path_for_sheet(path, sheet_name)
    if sheet_xml_path is None:
        return
    with zipfile.ZipFile(path, "r") as zin:
        entries = {n: zin.read(n) for n in zin.namelist()}
    if sheet_xml_path not in entries:
        return
    xml = entries[sheet_xml_path].decode("utf-8")
    if "<extLst>" in xml:
        return  # openpyxl unexpectedly kept its own -- don't write a duplicate
    xml = xml.replace("</worksheet>", ext_list_xml.decode("utf-8") + "</worksheet>")
    entries[sheet_xml_path] = xml.encode("utf-8")
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, data in entries.items():
            zout.writestr(name, data)


_CSV_PSEUDO_SHEET_NAME = "(CSV file)"


def list_sheet_names(path: str) -> list[str]:
    # A CSV has no sheets at all -- report one fixed pseudo-name so every
    # caller (the Client Setup sheet picker, in particular) can treat a
    # reference source's file as "always has at least one sheet to pick"
    # without a CSV-specific branch of its own.
    if path.lower().endswith(".csv"):
        return [_CSV_PSEUDO_SHEET_NAME]
    wb = openpyxl.load_workbook(path, read_only=True)
    try:
        return wb.sheetnames
    finally:
        wb.close()


def read_sheet_as_dataframe(path: str, sheet_name: str) -> pd.DataFrame:
    # sheet_name is meaningless for a CSV (there's only ever one "sheet")
    # and ignored here -- callers pass whatever list_sheet_names()
    # returned for this same path, i.e. _CSV_PSEUDO_SHEET_NAME.
    if path.lower().endswith(".csv"):
        with open(path, "rb") as f:
            return read_csv_bytes_robust(f.read())
    return pd.read_excel(path, sheet_name=sheet_name)


def dataframe_to_excel_bytes(df: pd.DataFrame, sheet_name: str = "Sheet1") -> bytes:
    """Serializes a DataFrame to real .xlsx bytes, for st.download_button --
    lets a user pull a list (refund/needs-review leads) into Excel to
    inspect without it ever touching disk on the server side."""
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
    return buffer.getvalue()


_CSV_ENCODINGS = ("utf-8-sig", "cp1252", "latin1")


def read_csv_bytes_robust(raw: bytes) -> pd.DataFrame:
    """Parse raw CSV bytes handling real-world export quirks: a UTF-8
    byte-order mark from Excel's own CSV export, Windows-1252 encoding
    from older export tools, and semicolon (or tab/pipe) delimiters from
    European-locale exports. Decoding bytes ourselves first, then sniffing
    the delimiter from the clean decoded text, avoids a pandas quirk where
    handing raw bytes + encoding= to read_csv(sep=None) can silently
    corrupt non-ASCII characters during delimiter detection even though
    the encoding itself is correct. Shared by every CSV entry point in
    this app (New Leads, reference/exclusion/TAL/suppression/dedupe
    sources, the Purchased Lead Report) so they all get the same handling.
    """
    text = None
    last_error: Exception | None = None
    for encoding in _CSV_ENCODINGS:
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError as exc:
            last_error = exc
    if text is None:
        raise last_error

    try:
        delimiter = csv.Sniffer().sniff(text[:4096], delimiters=",;\t|").delimiter
    except csv.Error:
        delimiter = ","

    return pd.read_csv(io.StringIO(text), sep=delimiter)


def read_leadfile(uploaded_file) -> pd.DataFrame:
    """Read an uploaded New Leads file — Excel or CSV — into a DataFrame."""
    name = getattr(uploaded_file, "name", "") or ""
    if not name.lower().endswith(".csv"):
        uploaded_file.seek(0)
        raw = uploaded_file.read()

        # Two independent problems, both producing the same "Unnamed: N"
        # symptom: (1) a workbook can have multiple sheets where the real
        # leads live on whichever sheet was active when the file was last
        # saved, not necessarily the first one by position — e.g. a pivot
        # summary sheet sitting before the real data sheet. wb.active is
        # openpyxl's read of that "last viewed" sheet, same as Excel shows
        # you on open; pd.read_excel's default sheet_name=0 has no such
        # concept and just takes the first by position, so it was reading
        # the wrong sheet entirely. (2) New Leads files, like Lead
        # Templates, sometimes carry a title/instruction row above the real
        # header, which a fixed header=0 reads as the header — detect the
        # true header row the same way find_header_row() does for target
        # sheets.
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True)
        try:
            sheet_name = wb.active.title
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20)))
            header_row = _detect_header_row_from_rows(rows, set(_ALL_KNOWN_HEADER_MARKERS))
        finally:
            wb.close()

        return pd.read_excel(io.BytesIO(raw), sheet_name=sheet_name, header=header_row - 1)

    uploaded_file.seek(0)
    return read_csv_bytes_robust(uploaded_file.read())


def _win_long_path(path: str) -> str:
    # Real client files often live in deeply nested OneDrive folders whose
    # full path is already close to Windows' 260-character MAX_PATH limit —
    # appending "_backup_<timestamp>" (or a "backup\" subfolder) pushes some
    # of them over it, which surfaces as a confusing WinError 3 ("cannot
    # find the path specified") even though every folder in the path is
    # real. The \\?\ prefix tells the Windows API to bypass that limit.
    if os.name != "nt":
        return path
    abs_path = os.path.abspath(path)
    if abs_path.startswith("\\\\?\\"):
        return abs_path
    if abs_path.startswith("\\\\"):
        return "\\\\?\\UNC\\" + abs_path[2:]
    return "\\\\?\\" + abs_path


def backup_file(path: str) -> str:
    source = Path(path)
    backup_dir = source.parent / "backup"
    os.makedirs(_win_long_path(str(backup_dir)), exist_ok=True)
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = backup_dir / f"{source.stem}_backup_{timestamp}{source.suffix}"
    shutil.copy2(_win_long_path(str(source)), _win_long_path(str(backup_path)))
    return str(backup_path)


def require_columns(df: pd.DataFrame, columns: list[str], file_label: str) -> None:
    missing = [c for c in columns if c not in df.columns]
    if missing:
        raise ValueError(
            f"'{file_label}' is missing expected column(s): {', '.join(missing)}. "
            f"Found columns: {', '.join(str(c) for c in df.columns)}"
        )


_FIELD_SYNONYMS = {
    "email": ["email", "emailaddress", "email address"],
    "first_name": ["firstname", "first name"],
    "last_name": ["lastname", "last name"],
    "company": ["company", "companyname", "company name"],
    "cid": ["cid"],
}


def _normalize_header_text(value) -> str:
    # Strips ALL non-alphanumeric characters (not just collapsing them to a
    # single space), so "Job Function", "Job_Function" and the fully
    # concatenated "jobfunction" all normalize to the same "jobfunction" —
    # real leadfiles frequently drop separators entirely, and requiring an
    # exact-ish phrase match left most non-core columns unmatched.
    import re
    return re.sub(r"[^a-z0-9]", "", str(value).strip().lower())


_NORMALIZED_FIELD_SYNONYMS = {
    attr: {_normalize_header_text(s) for s in synonyms} for attr, synonyms in _FIELD_SYNONYMS.items()
}


def _resolve_field_attr(header_norm: str) -> str | None:
    for attr, synonyms in _NORMALIZED_FIELD_SYNONYMS.items():
        if header_norm in synonyms:
            return attr
    return None


def guess_target_field_mapping(headers: list) -> dict[str, str]:
    """Best-guess header name for each of the 5 known lead fields, via the synonym table.

    Used only to seed a UI default — the caller should let the user confirm/override
    it and save an explicit mapping, since real target files (Accumulated Reports,
    Lead Templates) often use header text that doesn't match any synonym.
    """
    result: dict[str, str] = {}
    for header in headers:
        if header is None:
            continue
        attr = _resolve_field_attr(_normalize_header_text(header))
        if attr and attr not in result:
            result[attr] = header
    return result


_REASON_HEADER_NAMES = {_normalize_header_text(s) for s in ("reason", "refund reason")}

_ALL_KNOWN_HEADER_MARKERS = {_normalize_header_text(syn) for syns in _FIELD_SYNONYMS.values() for syn in syns}


_MIN_STRUCTURAL_HEADER_CELLS = 3


def _detect_header_row_from_rows(rows: list, markers: set[str]) -> int:
    """Shared two-tier header-row detection over a list of openpyxl rows
    (1-based row numbers preserved on each cell). See find_header_row for
    the detection strategy this implements.
    """
    for row in rows:
        for cell in row:
            if isinstance(cell.value, str) and _normalize_header_text(cell.value) in markers:
                return cell.row

    best_offset, best_count = None, 0
    for offset, row in enumerate(rows):
        non_empty = sum(1 for cell in row if cell.value is not None and str(cell.value).strip() != "")
        if non_empty >= _MIN_STRUCTURAL_HEADER_CELLS and non_empty > best_count:
            best_offset, best_count = offset, non_empty
    if best_offset is not None:
        return best_offset + 1

    return 1


def find_header_row(path: str, sheet_name: str, expected_headers: list | None = None,
                     max_scan_rows: int = 20) -> int:
    """Detect which row holds the real column headers.

    Some Lead Templates have title/instruction rows above the actual header
    row, so the headers don't start at row 1. Two-tier detection:

    1. Scan the first `max_scan_rows` rows for a cell matching a known marker
       (the client's saved column mapping if given, else the generic
       email/first name/last name/company/cid synonyms) and return that
       row's 1-based index — most reliable when the header text is
       recognizable.
    2. If no marker matches (e.g. the template uses non-standard header text
       with no saved mapping yet to compare against), fall back to a
       structural heuristic: among rows with at least
       `_MIN_STRUCTURAL_HEADER_CELLS` non-empty cells, pick the one with the
       MOST non-empty cells (earliest row wins on a tie) — real templates
       often carry annotation/instruction rows above the header that also
       have a handful of scattered notes in them, so "first row past a
       threshold" alone is not reliable; the true header row is reliably
       the densest one, since it names every column.

    Falls back to row 1 if neither tier finds anything, preserving the
    previous fixed-row-1 assumption.
    """
    markers = ({_normalize_header_text(h) for h in expected_headers if h}
               if expected_headers else set(_ALL_KNOWN_HEADER_MARKERS))

    wb = openpyxl.load_workbook(path, read_only=True)
    try:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=1, max_row=min(ws.max_row, max_scan_rows)))
        return _detect_header_row_from_rows(rows, markers)
    finally:
        wb.close()


def read_sheet_headers(path: str, sheet_name: str, header_row: int = 1) -> list:
    wb = openpyxl.load_workbook(path, read_only=True)
    try:
        ws = wb[sheet_name]
        # ws[header_row] (openpyxl's __getitem__) raises IndexError on a
        # genuinely empty sheet even though max_row reports 1 — iter_rows
        # doesn't have that problem, so use it instead.
        rows = list(ws.iter_rows(min_row=header_row, max_row=header_row))
        if not rows:
            return []
        return [cell.value for cell in rows[0]]
    finally:
        wb.close()


def route_leads_by_cid(
    leads_df: pd.DataFrame, cid_column: str, tabs: list[LeadTemplateTab], default_file_path: str = "",
) -> tuple[dict[tuple[str, str], pd.DataFrame], pd.DataFrame]:
    """Split leads across a Lead Template's tabs by CID.

    Tabs are checked in order; a lead is assigned to the first tab whose
    `cids` list contains its CID (mutually exclusive — a lead never lands in
    more than one tab). A tab with an empty `cids` list is never matched (a
    tab always needs its CIDs explicitly configured). Leads that match no
    tab at all are returned separately as "unmatched" rather than dropped.

    Returns ((file_path, sheet_name) -> matched leads for tabs with at least
    one match, unmatched leads). A tab's own file_path is used when set —
    some CID groups go to an entirely different workbook, not just another
    tab in the same one — otherwise falling back to default_file_path (the
    client's shared Lead Template path).
    """
    remaining = leads_df
    groups: dict[tuple[str, str], pd.DataFrame] = {}
    for tab in tabs:
        if not tab.cids or remaining.empty:
            continue
        mask = remaining[cid_column].astype(str).str.strip().isin(tab.cids)
        matched = remaining[mask]
        remaining = remaining[~mask]
        if not matched.empty:
            key = (tab.file_path or default_file_path, tab.sheet_name)
            groups[key] = pd.concat([groups[key], matched]) if key in groups else matched
    return groups, remaining


def _find_last_data_row(ws, first_data_row: int, headers: list) -> int | None:
    """Return the last row >= first_data_row with a real value in any of the
    given header's columns, or None if every such row is empty.

    ws.max_row reflects the sheet's whole used range, which includes cells
    that only ever had formatting applied — real templates are often
    bulk-preformatted thousands of rows past the actual data, which made
    ws.max_row alone report a row far below the true last lead. Scanning
    actual cell values is the only reliable way to find where leads end.
    """
    if ws.max_row < first_data_row:
        return None
    last = None
    num_cols = len(headers)
    # values_only=True skips constructing full Cell wrapper objects (style
    # refs, comments, hyperlinks) for every one of potentially tens of
    # thousands of rows — this scan only needs the raw values.
    for offset, row in enumerate(ws.iter_rows(min_row=first_data_row, max_row=ws.max_row, values_only=True)):
        if any(v is not None for v in row[:num_cols]):
            last = first_data_row + offset
    return last


def _cell_has_fill_color(cell, hex_color: str) -> bool:
    fill = cell.fill
    if fill is None or fill.fill_type != "solid":
        return False
    rgb = getattr(fill.fgColor, "rgb", None)
    # openpyxl ARGB strings carry an alpha prefix (e.g. "FFC6E0B4"), so
    # compare by suffix rather than exact equality.
    return isinstance(rgb, str) and rgb.upper().endswith(hex_color.upper())


_CONTAINMENT_MIN_LEN = 4
_FUZZY_MATCH_THRESHOLD = 88

# Known passthrough-column synonym groups: real leadfiles and Lead Templates
# use genuinely different words for the same field (not just typos/word
# order, which the fuzzy tier below already catches) — e.g. "Company Size"
# vs "Employee Size" — so no amount of string-similarity scoring will ever
# match them; "company" and "employee" just aren't similar strings. Each
# inner set is normalized synonyms for one field; add more groups here as
# further real-world mismatches turn up.
_PASSTHROUGH_SYNONYM_GROUPS: list[set[str]] = [
    {_normalize_header_text(s) for s in (
        "company size", "employee size", "employee count", "number of employees",
        "headcount", "company headcount", "employee size range", "company size range",
        "no of employees",
    )},
]
_PASSTHROUGH_SYNONYM_GROUP_BY_HEADER: dict[str, int] = {
    header: group_idx for group_idx, group in enumerate(_PASSTHROUGH_SYNONYM_GROUPS) for header in group
}


def _find_passthrough_lead_column(header_norm: str, lead_headers_norm: dict[str, str]) -> str | None:
    """Best-effort match of a target header to a leadfile column, for the
    "everything else" passthrough columns (beyond the 5 explicitly-mapped
    roles). Real leadfiles vary in ways an exact match can't anticipate —
    export tools append suffixes ("MarketSegmentReferential" for "Market
    Segment") or contract phrases ("IAMAReferential" for "I am a"). Tried in
    order, most to least confident:

    1. Exact match on the fully-stripped normalized text (handles
       "Job Function" / "jobfunction").
    2. Known synonym group (handles genuinely different wording for the same
       field, like "Company Size" vs "Employee Size" — see
       _PASSTHROUGH_SYNONYM_GROUPS above).
    3. Containment: one normalized string is fully contained in the other
       (handles suffix/prefix noise like "Referential") — guarded by a
       minimum length so short strings ("cid") don't swallow unrelated
       columns.
    4. Fuzzy similarity (rapidfuzz) above a high threshold, for typos and
       reordered words.

    A tier is only used if exactly one leadfile column qualifies — wiring
    the wrong column into a client's real report is worse than leaving a
    cell blank, so ties are left unmatched rather than guessed.
    """
    if header_norm in lead_headers_norm:
        return lead_headers_norm[header_norm]

    group_idx = _PASSTHROUGH_SYNONYM_GROUP_BY_HEADER.get(header_norm)
    if group_idx is not None:
        candidates = [
            orig for norm, orig in lead_headers_norm.items()
            if _PASSTHROUGH_SYNONYM_GROUP_BY_HEADER.get(norm) == group_idx
        ]
        if len(candidates) == 1:
            return candidates[0]

    if len(header_norm) >= _CONTAINMENT_MIN_LEN:
        candidates = [
            orig for norm, orig in lead_headers_norm.items()
            if len(norm) >= _CONTAINMENT_MIN_LEN and (norm in header_norm or header_norm in norm)
        ]
        if len(candidates) == 1:
            return candidates[0]

    from rapidfuzz import fuzz
    scored = sorted(
        ((fuzz.ratio(header_norm, norm), orig) for norm, orig in lead_headers_norm.items()),
        key=lambda pair: pair[0], reverse=True,
    )
    if scored and scored[0][0] >= _FUZZY_MATCH_THRESHOLD and (len(scored) == 1 or scored[0][0] > scored[1][0]):
        return scored[0][1]

    return None


def append_leads(
    accumulated_path: str,
    tab_name: str,
    leads_df: pd.DataFrame,
    field_mapping: FieldMapping,
    run_date,
    reasons: dict[int, str] | None = None,
    target_field_mapping: FieldMapping | None = None,
    header_row: int = 1,
    clear_existing: bool = False,
    highlight_fill: str | None = None,
) -> list[str]:
    _original_external_links = _read_external_link_parts(accumulated_path)
    _original_ext_list = _read_worksheet_ext_list(accumulated_path, tab_name)

    wb = openpyxl.load_workbook(accumulated_path)
    ws = wb[tab_name]

    headers = [cell.value for cell in ws[header_row]]
    lead_headers_norm = {_normalize_header_text(h): h for h in leads_df.columns}

    target_role_by_header: dict[str, str] = {}
    if target_field_mapping is not None:
        for attr in ("email", "first_name", "last_name", "company", "cid"):
            target_header = getattr(target_field_mapping, attr, "")
            if target_header:
                target_role_by_header[_normalize_header_text(target_header)] = attr

    has_reason_column = any(
        h is not None and _normalize_header_text(h) in _REASON_HEADER_NAMES for h in headers
    )
    if reasons and not has_reason_column:
        reason_col_idx = len(headers) + 1
        ws.cell(row=header_row, column=reason_col_idx, value="Refund Reason")
        headers.append("Refund Reason")

    first_data_row = header_row + 1
    formula_template: dict[str, tuple[str, str]] = {}
    if ws.max_row >= first_data_row:
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=first_data_row, column=col_idx)
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formula_template[header] = (cell.value, cell.coordinate)

    # clear_existing wipes old data rows (e.g. a Lead Report re-sent fresh
    # each period rather than accumulated) — capture the formatting from
    # the row about to be deleted first, since there'll be nothing left to
    # sample it from afterward.
    cleared_styles: dict[int, tuple] | None = None
    if clear_existing and ws.max_row >= first_data_row:
        cleared_styles = {}
        for col_idx in range(1, len(headers) + 1):
            src = ws.cell(row=first_data_row, column=col_idx)
            cleared_styles[col_idx] = (src.font, src.fill, src.border, src.alignment, src.number_format)
        ws.delete_rows(first_data_row, ws.max_row - first_data_row + 1)

    last_data_row = _find_last_data_row(ws, first_data_row, headers)
    has_existing_leads = last_data_row is not None
    style_template_row = (
        last_data_row if has_existing_leads
        else (first_data_row if ws.max_row >= first_data_row else None)
    )
    column_styles: dict[int, tuple] = {}
    if cleared_styles is not None:
        column_styles = cleared_styles
    elif style_template_row is not None:
        for col_idx in range(1, len(headers) + 1):
            src = ws.cell(row=style_template_row, column=col_idx)
            column_styles[col_idx] = (src.font, src.fill, src.border, src.alignment, src.number_format)

    # Which lead column (if any) feeds each header only depends on the
    # header/column identity, never on a specific row — resolve it once
    # per column rather than once per (row, column) pair.
    column_source: dict[int, str | None] = {}
    unmatched_passthrough_headers: list[str] = []
    for col_idx, header in enumerate(headers, start=1):
        if header is None:
            continue
        header_norm = _normalize_header_text(header)
        if header_norm in ("date", "comment", "status") or header in formula_template or header_norm in _REASON_HEADER_NAMES:
            continue
        if header_norm in target_role_by_header:
            column_source[col_idx] = getattr(field_mapping, target_role_by_header[header_norm])
            continue
        attr = _resolve_field_attr(header_norm)
        if attr:
            column_source[col_idx] = getattr(field_mapping, attr)
            continue
        source_col = _find_passthrough_lead_column(header_norm, lead_headers_norm)
        column_source[col_idx] = source_col
        if source_col is None:
            unmatched_passthrough_headers.append(header)

    next_row = first_data_row if not has_existing_leads else last_data_row + 1
    for row_offset, (idx, lead_row) in enumerate(leads_df.iterrows()):
        excel_row = next_row + row_offset
        for col_idx, header in enumerate(headers, start=1):
            if header is None:
                continue
            header_norm = _normalize_header_text(header)
            cell = ws.cell(row=excel_row, column=col_idx)
            if col_idx in column_styles:
                font, fill, border, alignment, number_format = column_styles[col_idx]
                cell.font, cell.fill, cell.border, cell.alignment, cell.number_format = (
                    copy(font), copy(fill), copy(border), copy(alignment), number_format
                )
            # Captured before assigning cell.value below — openpyxl itself
            # overwrites a "General" cell's number_format the moment a
            # date/datetime value is assigned to it, so checking *after*
            # assignment would always see openpyxl's own default format
            # instead of the "General" it actually started from.
            was_general_format = cell.number_format == "General"

            if header_norm == "date":
                cell.value = run_date
                if was_general_format and isinstance(cell.value, (datetime.date, datetime.datetime)):
                    cell.number_format = "dd-mmm-yy"
            elif header in formula_template:
                formula, origin_ref = formula_template[header]
                col_letter = get_column_letter(col_idx)
                cell.value = Translator(formula, origin=origin_ref).translate_formula(f"{col_letter}{excel_row}")
            elif header_norm in ("comment", "status"):
                cell.value = None
            elif header_norm in _REASON_HEADER_NAMES:
                cell.value = (reasons or {}).get(idx, "")
            else:
                source_col = column_source.get(col_idx)
                cell.value = lead_row.get(source_col, "") if source_col is not None else None
                # A real date/datetime value written into a "General"-formatted
                # cell displays as a raw serial number and Excel's date filter
                # can't group it — give it an explicit date format so it shows
                # and filters like a real date instead.
                if was_general_format and isinstance(cell.value, (datetime.date, datetime.datetime)):
                    cell.number_format = "mm/dd/yyyy"

    if highlight_fill and not leads_df.empty:
        # Only ever one batch highlighted at a time — clear this same color
        # from whatever rows existed before this run (an earlier run's
        # highlight), then apply it fresh to the rows this run just added.
        for row in range(first_data_row, next_row):
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row, column=col_idx)
                if _cell_has_fill_color(cell, highlight_fill):
                    cell.fill = PatternFill(fill_type=None)
        new_fill = PatternFill(start_color=highlight_fill, end_color=highlight_fill, fill_type="solid")
        for row in range(next_row, next_row + len(leads_df)):
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row, column=col_idx).fill = new_fill

    # openpyxl doesn't keep an Excel Table's declared range in sync with
    # delete_rows() or new cell writes on its own -- after clear_existing
    # deletes rows (or leads just get appended past the old range), a
    # table's `ref` goes stale relative to the sheet's real data footprint
    # (observed: a table still claiming rows 2-99 after clearing left only
    # 2 real data rows). Some downstream ingestion platforms read a sheet
    # via its declared table range rather than a raw cell scan, and see
    # that mismatch as "no data in the file" even though the cells are
    # populated. Resize every table anchored at this header row to match.
    final_last_row = next_row + len(leads_df) - 1
    for table in ws.tables.values():
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        if min_row <= header_row <= max_row:
            table.ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{final_last_row}"

    wb.save(accumulated_path)
    wb.close()

    if _original_external_links:
        _restore_external_link_parts(accumulated_path, _original_external_links)
    if _original_ext_list:
        _restore_worksheet_ext_list(accumulated_path, tab_name, _original_ext_list)

    return unmatched_passthrough_headers


def detect_cids_from_pacing_overview(
    accumulated_path: str,
    sheet_name: str = "Pacing Overview",
) -> list[tuple[str, str]]:
    wb = openpyxl.load_workbook(accumulated_path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"'{accumulated_path}' has no sheet named '{sheet_name}'")
        ws = wb[sheet_name]

        header_row_idx = None
        cid_col_idx = None
        campaign_col_idx = None
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20)):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip().lower() == "cid":
                    header_row_idx = cell.row
                    cid_col_idx = cell.column
            if header_row_idx is not None:
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.strip().lower() in (
                        "campaign segment", "campaign name", "campaign"
                    ):
                        campaign_col_idx = cell.column
                break

        if header_row_idx is None or cid_col_idx is None:
            raise ValueError(f"Could not find a 'CID' column in '{accumulated_path}' [{sheet_name}]")

        pairs: list[tuple[str, str]] = []
        seen: set[str] = set()
        for row in ws.iter_rows(min_row=header_row_idx + 1, max_row=ws.max_row):
            cid_cell = row[cid_col_idx - 1]
            if cid_cell.value is None or str(cid_cell.value).strip() == "":
                break
            cid = str(cid_cell.value).strip()
            if cid.lower() in ("grand total", "total"):
                break
            if cid in seen:
                continue
            seen.add(cid)
            campaign = ""
            if campaign_col_idx is not None:
                campaign_cell = row[campaign_col_idx - 1]
                if campaign_cell.value is not None:
                    campaign = str(campaign_cell.value).strip()
            pairs.append((cid, campaign or cid))

        return pairs
    finally:
        wb.close()


def _format_pacing_header(value) -> str:
    # Date column headers ("19-Aug") are stored as real datetime values, not
    # text — str()'ing one directly gives a full "2026-08-19 00:00:00"
    # timestamp instead of the short date the sheet actually displays.
    if isinstance(value, (datetime.datetime, datetime.date)):
        return value.strftime("%d-%b")
    return str(value).strip()


def _format_pacing_value(header: str, value):
    if value is None:
        return ""
    if _normalize_header_text(header) == "pacing" and isinstance(value, (int, float)):
        return f"{value * 100:.0f}%"
    return value


def read_pacing_overview_table(accumulated_path: str, sheet_name: str = "Pacing Overview") -> pd.DataFrame:
    """Read the Pacing Overview sheet as a full table (every column, not
    just CID/Campaign), for embedding as a native table in a Jira summary
    comment. Reuses the same header-row rule as detect_cids_from_pacing_overview
    (header row = the row with a "CID" cell). Stops at the first blank CID;
    a trailing "Grand Total"/"Total" row is kept (as the last row) rather
    than dropped, since it's a real, meaningful summary row in the source
    sheet. The "Pacing" column is formatted as a percentage and date column
    headers are shortened to a plain date, matching how the sheet actually
    displays them. A column the sheet itself has hidden (e.g. an old or
    not-yet-active date the client collapsed to reduce clutter) is skipped
    — openpyxl reads a hidden column's values same as any other, so without
    this check the Jira table would show a date the sheet doesn't visibly
    have. Reading column visibility requires a non-read-only load — that
    metadata isn't available in read_only mode.
    """
    wb = openpyxl.load_workbook(accumulated_path, read_only=False, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"'{accumulated_path}' has no sheet named '{sheet_name}'")
        ws = wb[sheet_name]

        header_row_idx = None
        cid_col_idx = None
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20)):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip().lower() == "cid":
                    header_row_idx = cell.row
                    cid_col_idx = cell.column
            if header_row_idx is not None:
                break

        if header_row_idx is None or cid_col_idx is None:
            raise ValueError(f"Could not find a 'CID' column in '{accumulated_path}' [{sheet_name}]")

        header_cells = next(ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx))
        col_indices = [
            c.column for c in header_cells
            if c.value is not None and str(c.value).strip() != ""
            and not ws.column_dimensions[get_column_letter(c.column)].hidden
        ]
        headers = [_format_pacing_header(ws.cell(row=header_row_idx, column=c).value) for c in col_indices]

        def _row_record(row) -> dict:
            return {
                header: _format_pacing_value(header, row[col - 1].value)
                for header, col in zip(headers, col_indices)
            }

        records: list[dict] = []
        for row in ws.iter_rows(min_row=header_row_idx + 1, max_row=ws.max_row):
            cid_cell = row[cid_col_idx - 1]
            if cid_cell.value is None or str(cid_cell.value).strip() == "":
                break
            records.append(_row_record(row))
            if str(cid_cell.value).strip().lower() in ("grand total", "total"):
                break

        return pd.DataFrame.from_records(records, columns=headers)
    finally:
        wb.close()
